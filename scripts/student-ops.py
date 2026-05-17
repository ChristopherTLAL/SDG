#!/usr/bin/env python3
"""
student-ops.py — vault student lifecycle CLI.

Subcommands:
  onboard <cid> [--advisor X] [--intake YEAR] [--name NAME]
     Build new vault folder from ERP cid. Pulls 合同明细 from full-history ERP.
     Defaults derived from ERP if not specified.

  transfer <student> --to <advisor> [--reason TEXT]
     Replace 中期顾问 with new advisor, insert 转案记录 in body.

  coadvise <student> --add <advisor>
     Append advisor to 中期顾问 array (convert scalar to array if needed).
     For 共带 cases.

  status <student> --to <stage>
     Update 当前进度 (one of: 中期在途/后期在途/已结案/退费/需对接).

Each subcommand writes vault .md, then prints a suggested git commit command
(does NOT auto-commit, you should review the diff).

Examples:
  scripts/student-ops.py onboard P24AAB+VEBtxmUc98028
  scripts/student-ops.py onboard P24AAB+VEBtxmUc98028 --advisor 王姝琰 --intake 2028
  scripts/student-ops.py transfer 颜佳嫣 --to 钟婷婷 --reason "古淑婷转过来"
  scripts/student-ops.py coadvise 张诗瑜 --add 高幸玲
  scripts/student-ops.py status 颜佳嫣 --to 退费
"""
import sys
import argparse
import re
import importlib.util
from pathlib import Path
from datetime import datetime
from collections import defaultdict
import openpyxl

VAULT_STUDENT = Path('/Users/shijie/Library/CloudStorage/OneDrive-Personal/Obsidian/规划看板/01_Student')
SDG_HTML = Path('/Users/shijie/Code/sdg-html')

# Import helpers from import-signings.py — must clear our argv first or its CLI
# parser will try to interpret e.g. "onboard" as an xlsx path and fail.
_orig_argv = sys.argv
sys.argv = [sys.argv[0]]
spec = importlib.util.spec_from_file_location('imp', SDG_HTML / 'scripts/import-signings.py')
imp = importlib.util.module_from_spec(spec)
spec.loader.exec_module(imp)
sys.argv = _orig_argv

TODAY = datetime.now().strftime('%Y-%m-%d')


# ── Helpers ──────────────────────────────────────────────────────────────

def find_folder(name: str) -> Path:
    """Resolve student folder by exact name match in 01_Student/."""
    direct = VAULT_STUDENT / name
    if direct.is_dir():
        return direct
    raise SystemExit(
        f'❌ No folder found for "{name}". '
        f'Try the EXACT folder name (incl. "(中期顾问)" suffix if any).'
    )


def parse_advisors_field(yaml_val: str) -> list:
    """Parse 中期顾问 value (scalar or array) → list of names."""
    v = yaml_val.strip()
    if not v:
        return []
    if v.startswith('['):
        items = [s.strip().strip('"').strip("'") for s in v.strip('[]').split(',')]
        return [x for x in items if x]
    return [v.strip('"').strip("'")]


def read_erp_contracts_for_cid(cid: str):
    """Pull all contract rows for a cid from the full-history ERP file.
    Returns (contracts_list, customer_name, primary_intake, primary_mid_advisor).
    Primary = from latest 留学申请=是 contract.
    """
    wb = openpyxl.load_workbook(imp.XLSX, read_only=True, data_only=True)
    ws = wb['签约明细']
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    COL = {h: i for i, h in enumerate(header) if h}
    contracts = []
    customer_name = None
    primary_intake = None
    primary_mid = None
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not (r and r[COL['客户ID']] == cid):
            continue
        signed = imp.parse_date(r[COL['签约时间']])
        if not signed:
            continue
        customer_name = customer_name or r[COL['客户姓名']]
        tmpl = r[COL['合同模板名称']] or ''
        is_app = imp.is_yes(r[COL['是否包含留学申请']])
        contracts.append({
            'contract_id': r[COL['合同编号']],
            'customer_id': cid,
            'customer_name': r[COL['客户姓名']],
            'signed_at': signed,
            'signed_amount': r[COL['签约金额（实时）']] or 0,
            'template_name': tmpl,
            'category': imp.category(tmpl),
            'status': r[COL['合同状态']] or '',
        })
        # primary intake / mid from latest 留学申请 contract
        if is_app:
            if r[COL['入学年']] and not primary_intake:
                primary_intake = r[COL['入学年']]
            if r[COL['中期顾问']] and not primary_mid:
                primary_mid = r[COL['中期顾问']]
    contracts.sort(key=lambda c: c['signed_at'], reverse=True)
    return contracts, customer_name, primary_intake, primary_mid


def print_commit_hint(folder_name: str, msg: str):
    print(f'\n📋 建议 commit:')
    print(f'   cd "$VAULT" && git add 01_Student/"{folder_name}" \\\n     && git commit -m "{msg}"')


# ── Subcommand: onboard ──────────────────────────────────────────────────

def cmd_onboard(args):
    contracts, erp_name, erp_intake, erp_mid = read_erp_contracts_for_cid(args.cid)
    if not contracts:
        raise SystemExit(f'❌ cid {args.cid} not found in ERP {imp.XLSX.name}')

    name = args.name or erp_name
    advisor = args.advisor or erp_mid
    intake = args.intake or (str(erp_intake) if erp_intake else None)

    if not advisor:
        raise SystemExit(
            f'❌ No 中期顾问 found in ERP 留学合同 for {args.cid}. Pass --advisor explicitly.'
        )
    if not intake:
        raise SystemExit(
            f'❌ No 入学年 found in ERP 留学合同 for {args.cid}. Pass --intake explicitly.'
        )

    target_folder = name
    existing = {p.name for p in VAULT_STUDENT.iterdir() if p.is_dir()}
    if target_folder in existing:
        target_folder = f'{name} ({advisor})'
        print(f'⚠ Folder "{name}" already exists. Using suffix → "{target_folder}"')

    ok = imp.create_new_student_folder(target_folder, advisor, intake, args.cid, contracts)
    if ok:
        print(f'✓ Built vault/{target_folder}/ (中期={advisor}, intake={intake}F, 合同={len(contracts)})')
        print_commit_hint(target_folder, f'onboard {target_folder} ({advisor}, {intake}F)')
    else:
        print(f'· {target_folder} already exists, no change made.')


# ── Subcommand: transfer ─────────────────────────────────────────────────

def cmd_transfer(args):
    folder = find_folder(args.student)
    md = folder / f'{folder.name}.md'
    text = md.read_text(encoding='utf-8')

    m = re.search(r'^中期顾问:[ \t]*(.*?)$', text, re.MULTILINE)
    if not m:
        raise SystemExit(f'❌ No 中期顾问 line in {md}')
    old_advisors = parse_advisors_field(m.group(1))
    old_label = ', '.join(old_advisors) if old_advisors else '(空)'

    if args.to in old_advisors and len(old_advisors) == 1:
        print(f'· {args.to} already is the 中期顾问 of {folder.name}, nothing to do.')
        return

    new_text = text[:m.start()] + f'中期顾问: {args.to}' + text[m.end():]

    reason = args.reason or '按用户确认'
    note = (
        f'\n## 转案记录\n\n'
        f'- **{TODAY}**: 中期顾问 {old_label} → **{args.to}**'
        f'（{reason}；ERP 中期顾问字段待运营更新）\n\n'
    )
    h1 = re.search(r'^# .*?$\n', new_text, re.MULTILINE)
    if h1:
        new_text = new_text[:h1.end()] + note + new_text[h1.end():]
    else:
        new_text = new_text + note

    md.write_text(new_text, encoding='utf-8')
    print(f'✓ Transferred {folder.name}: 中期顾问 {old_label} → {args.to}')
    print(f'  + 转案记录 inserted in body')
    print_commit_hint(folder.name, f'{folder.name}: transfer 中期 {old_label} → {args.to}')


# ── Subcommand: coadvise ─────────────────────────────────────────────────

def cmd_coadvise(args):
    folder = find_folder(args.student)
    md = folder / f'{folder.name}.md'
    text = md.read_text(encoding='utf-8')

    m = re.search(r'^中期顾问:[ \t]*(.*?)$', text, re.MULTILINE)
    if not m:
        raise SystemExit(f'❌ No 中期顾问 line in {md}')
    existing = parse_advisors_field(m.group(1))

    if args.add in existing:
        print(f'· {args.add} already in 中期顾问 list ({existing}), nothing to do.')
        return

    new_list = existing + [args.add]
    new_line = f'中期顾问: [{", ".join(new_list)}]'
    new_text = text[:m.start()] + new_line + text[m.end():]
    md.write_text(new_text, encoding='utf-8')
    print(f'✓ Coadvise: {folder.name} 中期顾问 {existing} → {new_list}')
    print_commit_hint(folder.name, f'{folder.name}: coadvise + {args.add}')


# ── Subcommand: status ───────────────────────────────────────────────────

VALID_STAGES = {'中期在途', '后期在途', '已结案', '退费', '需对接'}

def cmd_status(args):
    if args.to not in VALID_STAGES:
        raise SystemExit(f'❌ --to must be one of: {sorted(VALID_STAGES)}')
    folder = find_folder(args.student)
    md = folder / f'{folder.name}.md'
    text = md.read_text(encoding='utf-8')
    new_text, n = re.subn(
        r'^当前进度:[ \t]*(.*?)$',
        f'当前进度: {args.to}',
        text, count=1, flags=re.MULTILINE,
    )
    if n == 0:
        raise SystemExit(f'❌ No 当前进度 line in {md}')
    if new_text == text:
        print(f'· {folder.name} 当前进度 already = {args.to}, nothing to do.')
        return
    md.write_text(new_text, encoding='utf-8')
    print(f'✓ {folder.name}: 当前进度 → {args.to}')
    print_commit_hint(folder.name, f'{folder.name}: status → {args.to}')


# ── Main ─────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawTextHelpFormatter,
    )
    sub = parser.add_subparsers(dest='subcmd', required=True, metavar='<subcmd>')

    p_on = sub.add_parser('onboard', help='Build new vault folder from ERP cid')
    p_on.add_argument('cid', help='ERP 客户ID (e.g. P24AAB+VEBtxmUc98028)')
    p_on.add_argument('--name', help='Override 客户姓名 (else use ERP)')
    p_on.add_argument('--advisor', help='Override 中期顾问 (else use ERP 留学合同)')
    p_on.add_argument('--intake', help='Override 入学年 (else use ERP, e.g. 2027)')
    p_on.set_defaults(func=cmd_onboard)

    p_tr = sub.add_parser('transfer', help='Replace 中期顾问 + add 转案记录')
    p_tr.add_argument('student', help='Vault folder name (exact)')
    p_tr.add_argument('--to', required=True, help='New 中期顾问')
    p_tr.add_argument('--reason', help='Reason note (default: 按用户确认)')
    p_tr.set_defaults(func=cmd_transfer)

    p_co = sub.add_parser('coadvise', help='Append second 中期顾问 (共带)')
    p_co.add_argument('student', help='Vault folder name (exact)')
    p_co.add_argument('--add', required=True, help='Additional 中期顾问 to append')
    p_co.set_defaults(func=cmd_coadvise)

    p_st = sub.add_parser('status', help='Update 当前进度')
    p_st.add_argument('student', help='Vault folder name (exact)')
    p_st.add_argument(
        '--to', required=True,
        help='One of: 中期在途 / 后期在途 / 已结案 / 退费 / 需对接',
    )
    p_st.set_defaults(func=cmd_status)

    args = parser.parse_args()
    args.func(args)


if __name__ == '__main__':
    main()
