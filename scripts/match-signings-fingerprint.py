#!/usr/bin/env python3
"""
match-signings-fingerprint.py — read-only fingerprint of signing-table ↔ vault.

Reads ~/Downloads/2024.6.1-2026.5.3客户签约明细.xlsx
Writes /tmp/signing-match-report.md

Helps user decide: which vault folders need alias mapping, which signing
customers should be onboarded, which 重名 cases need disambiguation.
"""
import openpyxl
import re
import sys
from pathlib import Path
from collections import Counter, defaultdict
from datetime import datetime

# Pick the latest 客户签约明细 export from ~/Downloads (sorted by mtime).
def _find_latest_signing_xlsx():
    candidates = sorted(
        (p for p in (Path.home() / 'Downloads').glob('*客户签约明细*.xlsx')
         if not p.name.startswith('~$')),  # skip Excel lock files
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError('No *客户签约明细*.xlsx found in ~/Downloads')
    return candidates[0]

XLSX = _find_latest_signing_xlsx()
VAULT = Path('/Users/shijie/Obsidian/规划看板/01_Student')
OUT = Path('/tmp/signing-match-report.md')

# Known aliases user has confirmed:
KNOWN_ALIASES = {
    'Kimi+Byran': ['LI KIMI', 'Bryan Wang'],
    'Sophia+Amy': ['LI XINYING', 'Sophia Wang'],
}


def read_signings():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb['签约明细']
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    COL = {h: i for i, h in enumerate(header) if h}
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and r[COL['客户姓名']]:
            rows.append(r)
    return rows, COL


def read_vault_students():
    """Returns list of {name, mid_advisor, alias_signers (from YAML if any), folder}"""
    students = []
    for p in VAULT.iterdir():
        if not p.is_dir() or p.name.startswith('_'):
            continue
        md = p / f'{p.name}.md'
        if not md.exists():
            continue
        text = md.read_text(encoding='utf-8', errors='ignore')
        # Extract YAML frontmatter
        m = re.match(r'^---\n(.*?)\n---', text, re.DOTALL)
        if not m:
            continue
        yaml_text = m.group(1)
        mid_advisor = None
        mid_advisors = []
        m2 = re.search(r'^中期顾问:\s*(\S.*)$', yaml_text, re.MULTILINE)
        if m2:
            v = m2.group(1).strip()
            mid_advisor = v
            mid_advisors = [v]
        m3 = re.search(r'^中期顾问:\s*\[([^\]]+)\]', yaml_text, re.MULTILINE)
        if m3:
            mid_advisors = [x.strip() for x in m3.group(1).split(',')]

        # Look for body memo mentioning parent / English name (heuristic for alias hints)
        memo_hints = []
        body = text[m.end():]
        # Common patterns: "父亲 X" / "母亲 X" / "妈妈 X" / 家长 X / 签约人 X
        for pat in [r'父亲[：:]?\s*([^\s，。、]+)', r'妈妈[：:]?\s*([^\s，。、]+)', r'母亲[：:]?\s*([^\s，。、]+)', r'家长[：:]?\s*([^\s，。、]+)']:
            for mm in re.finditer(pat, body):
                memo_hints.append(mm.group(1))
        students.append({
            'name': p.name,
            'mid_advisor': mid_advisor,
            'mid_advisors': mid_advisors,
            'memo_hints': memo_hints,
            'has_alias_signer': '合同签约人:' in yaml_text,
        })
    return students


def signing_advisor_for_name(rows, COL, name):
    """For a 客户姓名, return Counter of 中期顾问 from their signing rows."""
    advisors = Counter()
    for r in rows:
        if r[COL['客户姓名']] == name and r[COL['中期顾问']]:
            advisors[r[COL['中期顾问']]] += 1
    return advisors


def signing_summary_for_id(rows, COL, cust_id):
    matched = [r for r in rows if r[COL['客户ID']] == cust_id]
    total_amt = sum(r[COL['签约金额（实时）']] or 0 for r in matched)
    advisors = Counter(r[COL['中期顾问']] for r in matched if r[COL['中期顾问']])
    return {
        'count': len(matched),
        'total_amount': total_amt,
        'top_advisor': advisors.most_common(1)[0][0] if advisors else None,
        'is_application': any(r[COL['是否包含留学申请']] == '是' for r in matched),
    }


def main():
    print(f'Reading {XLSX}...')
    rows, COL = read_signings()
    print(f'  Total signed-contract rows: {len(rows)}')

    # Build name → ID list mapping (for 重名 detection)
    name_to_ids = defaultdict(set)
    id_to_rows = defaultdict(list)
    for r in rows:
        nm = r[COL['客户姓名']]
        cid = r[COL['客户ID']]
        if nm and cid:
            name_to_ids[nm].add(cid)
            id_to_rows[cid].append(r)

    print(f'  Unique 客户ID: {len(id_to_rows)}')
    print(f'  Unique 客户姓名: {len(name_to_ids)}')

    print(f'Reading vault {VAULT}...')
    vault = read_vault_students()
    print(f'  Vault students: {len(vault)}')

    # Stage 1: direct name match
    direct_matches = []   # (vault_student, customer_id)
    needs_alias = []      # vault students with 0 signing rows
    duplicates = []       # vault students whose name has >1 customer_id

    for vs in vault:
        ids = name_to_ids.get(vs['name'], set())
        if len(ids) == 0:
            needs_alias.append(vs)
        elif len(ids) == 1:
            direct_matches.append((vs, list(ids)[0]))
        else:
            duplicates.append((vs, list(ids)))

    # Stage 2: disambiguate duplicates by mid_advisor
    duplicate_resolutions = []  # (vault_student, [{id, summary, picked}])
    for vs, ids in duplicates:
        candidates = []
        for cid in ids:
            summ = signing_summary_for_id(rows, COL, cid)
            picked = vs['mid_advisor'] and summ['top_advisor'] == vs['mid_advisor']
            candidates.append({
                'id': cid,
                'summary': summ,
                'picked': picked,
            })
        # Auto-pick: if exactly one candidate's top_advisor matches vault mid_advisor
        picked_ids = [c['id'] for c in candidates if c['picked']]
        auto = picked_ids[0] if len(picked_ids) == 1 else None
        duplicate_resolutions.append({'vs': vs, 'candidates': candidates, 'auto_pick': auto})

    # Stage 3: alias suggestions for needs_alias
    alias_suggestions = []  # (vs, [candidate_signers])
    for vs in needs_alias:
        cands = []
        # Known alias?
        if vs['name'] in KNOWN_ALIASES:
            for nm in KNOWN_ALIASES[vs['name']]:
                ids = name_to_ids.get(nm, set())
                for cid in ids:
                    summ = signing_summary_for_id(rows, COL, cid)
                    cands.append({'name': nm, 'id': cid, 'summary': summ, 'source': '✅ 用户已确认'})
        # Split by '+' for combined-sibling folders
        if '+' in vs['name'] and vs['name'] not in KNOWN_ALIASES:
            for part in vs['name'].split('+'):
                part = part.strip()
                # Search signing for names containing this part (substring match)
                for sn in name_to_ids:
                    if part.lower() in sn.lower() and sn != part:
                        for cid in name_to_ids[sn]:
                            summ = signing_summary_for_id(rows, COL, cid)
                            cands.append({'name': sn, 'id': cid, 'summary': summ, 'source': f'split-by-+ "{part}"'})
        # Memo hints (父亲 / 妈妈 etc.)
        for hint in vs['memo_hints']:
            if hint in name_to_ids:
                for cid in name_to_ids[hint]:
                    summ = signing_summary_for_id(rows, COL, cid)
                    cands.append({'name': hint, 'id': cid, 'summary': summ, 'source': f'YAML 提到 "{hint}"'})
        alias_suggestions.append({'vs': vs, 'candidates': cands})

    # Reverse: high-value signing customers NOT in vault
    matched_ids = {cid for _, cid in direct_matches} | {r['auto_pick'] for r in duplicate_resolutions if r['auto_pick']}
    not_in_vault_summaries = []
    for cid, rs in id_to_rows.items():
        if cid in matched_ids:
            continue
        nm = rs[0][COL['客户姓名']]
        # Skip if this name is mentioned as alias
        if any(nm in (KNOWN_ALIASES.get(v['name'], []) + v['memo_hints']) for v in vault):
            continue
        total = sum(r[COL['签约金额（实时）']] or 0 for r in rs)
        is_app = any(r[COL['是否包含留学申请']] == '是' for r in rs)
        adv = Counter(r[COL['中期顾问']] for r in rs if r[COL['中期顾问']]).most_common(1)
        not_in_vault_summaries.append({
            'id': cid,
            'name': nm,
            'total': total,
            'count': len(rs),
            'top_advisor': adv[0][0] if adv else '-',
            'is_application': is_app,
        })
    # Top 20 留学 + top 10 非留学
    top_app = sorted([s for s in not_in_vault_summaries if s['is_application']], key=lambda x: -x['total'])[:20]
    top_nonapp = sorted([s for s in not_in_vault_summaries if not s['is_application']], key=lambda x: -x['total'])[:10]

    # Status / data-quality stats
    status_counts = Counter(r[COL['合同状态']] for r in rows if r[COL['合同状态']])
    is_app_counts = Counter(r[COL['是否包含留学申请']] for r in rows if r[COL['是否包含留学申请']])

    # Compose markdown
    md = []
    md.append('# 签约表 ↔ Vault 匹配指纹报告')
    md.append('')
    md.append(f'> 生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}')
    md.append(f'> 数据源：`{XLSX.name}`')
    md.append(f'> Vault 路径：`{VAULT}`')
    md.append('')

    md.append('## 总览')
    md.append('')
    md.append(f'- 签约表 **{len(rows)}** 行 / **{len(id_to_rows)}** 客户ID / **{len(name_to_ids)}** 客户姓名')
    md.append(f'- Vault **{len(vault)}** 学生')
    md.append(f'- ✅ 直接同名匹配 **{len(direct_matches)}**')
    md.append(f'- 🟡 重名干扰（vault 影响）**{len(duplicates)}**')
    md.append(f'- ⚠️ 0 直接命中 需 alias **{len(needs_alias)}**')
    md.append('')
    md.append('### 数据质量')
    md.append(f'- 合同状态分布：' + ' / '.join(f'{k} {v}' for k, v in status_counts.most_common()))
    md.append(f'- 留学申请：' + ' / '.join(f'{k} {v}' for k, v in is_app_counts.most_common()))
    md.append('')

    md.append('## 阶段 1：直接同名匹配（' + str(len(direct_matches)) + '，无需操作）')
    md.append('')
    md.append('<details><summary>展开 sample（前 10）</summary>')
    md.append('')
    md.append('| Vault | 客户ID | 合同数 | 总金额 | 顶部顾问 |')
    md.append('|:--|:--|:--|:--|:--|')
    for vs, cid in direct_matches[:10]:
        s = signing_summary_for_id(rows, COL, cid)
        md.append(f'| {vs["name"]} | `{cid}` | {s["count"]} | {s["total_amount"]:.0f} | {s["top_advisor"] or "-"} |')
    md.append('')
    md.append('</details>')
    md.append('')

    md.append('## 阶段 2：重名干扰（' + str(len(duplicates)) + '，需你确认）')
    md.append('')
    if not duplicates:
        md.append('（无）')
    else:
        for r in duplicate_resolutions:
            vs = r['vs']
            md.append(f'### {vs["name"]}')
            md.append(f'Vault 中期顾问：**{vs["mid_advisor"] or "未填"}**')
            md.append('')
            md.append('| 客户ID | 合同数 | 总金额 | 签约表顶部顾问 | 是否留学 | 自动判断 |')
            md.append('|:--|:--|:--|:--|:--|:--|')
            for c in r['candidates']:
                s = c['summary']
                pick = '✅ 选中' if c['picked'] and r['auto_pick'] == c['id'] else ('— 排除' if not c['picked'] else '🟡 候选')
                md.append(f'| `{c["id"]}` | {s["count"]} | {s["total_amount"]:.0f} | {s["top_advisor"] or "-"} | {"是" if s["is_application"] else "否"} | {pick} |')
            if r['auto_pick']:
                md.append(f'\n→ **自动判定：用 `{r["auto_pick"]}`**（顾问匹配）')
            else:
                md.append(f'\n→ ⚠️ **需你手动判断**（顾问不能消歧）')
            md.append('')

    md.append('## 阶段 3：0 直接命中 需 alias（' + str(len(needs_alias)) + '）')
    md.append('')
    for r in alias_suggestions:
        vs = r['vs']
        md.append(f'### {vs["name"]}')
        md.append(f'Vault 中期顾问：**{vs["mid_advisor"] or "未填"}**')
        if vs['memo_hints']:
            md.append(f'YAML 备注 hints：{", ".join(vs["memo_hints"])}')
        md.append('')
        if not r['candidates']:
            md.append('🚫 **无候选** — 可能是新客 / 老客（2024.6 之前）/ 别的产品线 / 内部测试 — 你看是否需要保留 vault folder')
            md.append('')
            continue
        md.append('| 候选签约人 | 客户ID | 合同数 | 金额 | 顾问 | 留学? | 来源 |')
        md.append('|:--|:--|:--|:--|:--|:--|:--|')
        seen = set()
        for c in r['candidates']:
            key = (c['name'], c['id'])
            if key in seen:
                continue
            seen.add(key)
            s = c['summary']
            md.append(f'| {c["name"]} | `{c["id"]}` | {s["count"]} | {s["total_amount"]:.0f} | {s["top_advisor"] or "-"} | {"是" if s["is_application"] else "否"} | {c["source"]} |')
        md.append('')

    md.append('## 反向：签约表里 vault 没的高价值客户')
    md.append('')
    md.append(f'### 留学申请类（top {len(top_app)} by 总金额）')
    md.append('')
    md.append('| 姓名 | 客户ID | 合同数 | 总金额 | 中期顾问 |')
    md.append('|:--|:--|:--|:--|:--|')
    for s in top_app:
        md.append(f'| {s["name"]} | `{s["id"]}` | {s["count"]} | {s["total"]:.0f} | {s["top_advisor"]} |')
    md.append('')

    md.append(f'### 非留学（培训/竞赛/班课）top {len(top_nonapp)} by 总金额')
    md.append('')
    md.append('| 姓名 | 客户ID | 合同数 | 总金额 | 中期顾问 |')
    md.append('|:--|:--|:--|:--|:--|')
    for s in top_nonapp:
        md.append(f'| {s["name"]} | `{s["id"]}` | {s["count"]} | {s["total"]:.0f} | {s["top_advisor"]} |')
    md.append('')

    md.append('## 你需要回我什么')
    md.append('')
    md.append('1. **阶段 2 重名**：自动判定不对的，告诉我用哪个客户ID')
    md.append('2. **阶段 3 alias**：每个 folder 对应哪些签约人（直接复制候选签约人姓名即可），不需要的可以删除 vault folder')
    md.append('3. **反向高价值**：哪些是漏建档的，要建 vault folder')
    md.append('')
    md.append('确认完，我写最终的 `import-signings.py`：')
    md.append('- 写 `客户ID` 进 vault YAML')
    md.append('- 重写 `合同明细`（去掉 预收余额，加 合同状态）')
    md.append('- 新建 Supabase `contracts` 表 + 灌全量数据')
    md.append('- 后续每周 idempotent re-run')

    OUT.write_text('\n'.join(md), encoding='utf-8')
    print(f'\n✅ Report written: {OUT}')
    print(f'   Lines: {len(md)}')
    print(f'\nQuick stats:')
    print(f'  Direct matches: {len(direct_matches)}')
    print(f'  Duplicate names (vault impacted): {len(duplicates)}')
    print(f'  Need alias: {len(needs_alias)}')
    print(f'  High-value signing not in vault (留学): {len(top_app)}')
    print(f'  High-value signing not in vault (非留学): {len(top_nonapp)}')


if __name__ == '__main__':
    main()
