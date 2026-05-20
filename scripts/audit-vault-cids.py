#!/usr/bin/env python3
"""
audit-vault-cids.py — 定期审计 vault ↔ ERP cid 绑定健康度 (read-only).

跑法:
  /usr/local/bin/python3 scripts/audit-vault-cids.py [path/to/*客户签约明细*.xlsx]
  默认读 ~/Downloads/_工作运营/客户签约明细-*.xlsx (全量历史)

报告分 6 节:
  A: ✓ Clean — vault 学生绑的 cid 在 ERP 找到，姓名一致
  B: ⚠ Name mismatch — vault folder 名跟 ERP 客户姓名 对不上 (可能 alias 或 误绑)
  C: ⚠ vault cid 在 ERP 找不到 — 可能错绑、或 ERP 历史数据不全 (我们只有近 3 年)
  D: ⚠ vault 学生没绑 cid (非私单) — 漏绑或老学生
  E: ℹ vault folder 绑多 cid — 通常是 sibling/alias (合理) 或 误绑
  F: ❌ ERP cid 没绑到任何 vault — orphan, by 中期顾问 分组待人工决策

跑这个不会改任何东西，纯诊断。
"""
import sys
import re
from pathlib import Path
from collections import defaultdict, Counter
import openpyxl

VAULT = Path('/Users/shijie/Obsidian/规划看板/01_Student')
DEFAULT_XLSX = Path.home() / 'Downloads' / '_工作运营' / '客户签约明细-2023-06至2026-05.xlsx'

if len(sys.argv) > 1:
    XLSX = Path(sys.argv[1]).expanduser()
else:
    XLSX = DEFAULT_XLSX
if not XLSX.exists():
    print(f'ERROR: {XLSX} not found. Pass a path or place file in ~/Downloads/_工作运营/', file=sys.stderr)
    sys.exit(1)

# In-house active 中期 advisors (Supabase 现有)
INHOUSE_ADVISORS = {'古淑婷', '张曌璐', '张逸旸', '徐祖蕴', '王世杰', '王姝琰',
                    '袁辰飞', '钟婷婷', '陆梦婕', '高幸玲'}

# Known ALIASES from import-signings.py (sibling-combined / alt-name signing)
ALIASES = {
    'Kimi+Byran':  ['LI KIMI', 'Bryan Wang'],
    'Sophia+Amy':  ['LI XINYING', 'Sophia Wang'],
    'Betty':       ['陈碧涵'],
}


def read_vault():
    students = []
    for p in sorted(VAULT.iterdir()):
        if not p.is_dir() or p.name.startswith('_'):
            continue
        md = p / f'{p.name}.md'
        if not md.exists():
            continue
        text = md.read_text(encoding='utf-8', errors='ignore')
        m = re.match(r'^---\s*\n(.*?)\n---', text, re.DOTALL)
        if not m:
            continue
        yaml = m.group(1)
        # parse 客户ID
        cids = []
        cm = re.search(r'^客户ID:[ \t]*\[(.*?)\]', yaml, re.MULTILINE)
        if cm:
            for tok in cm.group(1).split(','):
                t = tok.strip().strip('"').strip("'")
                if t: cids.append(t)
        # parse 中期顾问 (first if array)
        mid = ''
        midm = re.search(r'^中期顾问:[ \t]*(.*?)$', yaml, re.MULTILINE)
        if midm:
            v = midm.group(1).strip()
            if v.startswith('['):
                items = [s.strip().strip('"').strip("'") for s in v.strip('[]').split(',') if s.strip()]
                mid = items[0] if items else ''
            else:
                mid = v.strip('"').strip("'")
        # 私单?
        cm2 = re.search(r'^合同:[ \t]*(.*?)$', yaml, re.MULTILINE)
        is_priv = bool(cm2 and '私单' in cm2.group(1))
        # intake
        intm = re.search(r'^入学年份:[ \t]*(.*?)$', yaml, re.MULTILINE)
        intake = intm.group(1).strip() if intm else ''
        # stage
        stm = re.search(r'^当前进度:[ \t]*(.*?)$', yaml, re.MULTILINE)
        stage = stm.group(1).strip() if stm else ''
        students.append({
            'name': p.name,
            'cids': cids,
            'mid': mid,
            'is_private': is_priv,
            'intake': intake,
            'stage': stage,
        })
    return students


def read_erp():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb['签约明细']
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    COL = {h: i for i, h in enumerate(header) if h}
    cid_data = {}  # cid → aggregated info
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not (r and r[COL['客户姓名']] and r[COL['客户ID']]):
            continue
        cid = r[COL['客户ID']]
        nm = r[COL['客户姓名']]
        mid = r[COL['中期顾问']]
        is_app = r[COL['是否包含留学申请']] in ('是', True, 1, '1')
        status = str(r[COL['合同状态']] or '')
        intake = r[COL['入学年']]
        d = cid_data.setdefault(cid, {
            'names': set(), 'mid_set': set(),
            'app_mid_set': set(),  # only 留学申请合同的中期顾问
            'intake': None, 'app_intake': None,
            'has_app': False, 'has_active': False,
            'contract_count': 0,
        })
        d['names'].add(nm)
        d['contract_count'] += 1
        if mid: d['mid_set'].add(mid)
        if is_app:
            d['has_app'] = True
            if mid: d['app_mid_set'].add(mid)
            if intake: d['app_intake'] = intake
        if intake and not d['intake']: d['intake'] = intake
        if status and '终止' not in status and '退' not in status:
            d['has_active'] = True
    return cid_data


def main():
    print('📖 Loading vault + ERP')
    vault = read_vault()
    erp = read_erp()
    print(f'   vault: {len(vault)} students  |  ERP: {len(erp)} unique cids')

    # Build reverse lookup
    erp_name_to_cids = defaultdict(set)
    for cid, d in erp.items():
        for nm in d['names']:
            erp_name_to_cids[nm].add(cid)

    vault_cid_to_folder = {}
    for vs in vault:
        for cid in vs['cids']:
            vault_cid_to_folder[cid] = vs['name']

    # Aliases reverse
    alias_signer_to_folder = {}
    for folder, signers in ALIASES.items():
        for s in signers:
            alias_signer_to_folder[s] = folder

    # ── Section A/B/C/D/E ──
    sec_A_clean = []     # vault cid in ERP, name matches (or alias)
    sec_B_name_mismatch = []
    sec_C_cid_not_in_erp = []
    sec_D_no_cid = []
    sec_E_multi_cid = []

    for vs in vault:
        if vs['is_private']:
            continue  # 私单 不在审计内
        if not vs['cids']:
            sec_D_no_cid.append(vs)
            continue
        if len(vs['cids']) > 1:
            sec_E_multi_cid.append(vs)
        for cid in vs['cids']:
            if cid not in erp:
                sec_C_cid_not_in_erp.append((vs, cid))
                continue
            erp_names = erp[cid]['names']
            # ok if vault name in erp names, OR (vault has alias entry and erp name is in alias signers)
            ok = vs['name'] in erp_names
            if not ok and vs['name'] in ALIASES:
                ok = any(s in erp_names for s in ALIASES[vs['name']])
            if ok:
                sec_A_clean.append((vs, cid))
            else:
                sec_B_name_mismatch.append((vs, cid, sorted(erp_names)))

    # ── Section F: ERP cid not bound to any vault ──
    sec_F_orphan = []
    for cid, d in erp.items():
        if cid in vault_cid_to_folder:
            continue
        # Also check if ANY of the names is an alias signer (those should match)
        if any(nm in alias_signer_to_folder and alias_signer_to_folder[nm] in {vs['name'] for vs in vault} for nm in d['names']):
            # An aliased signer's cid — already represented via aliasing. Skip.
            continue
        sec_F_orphan.append((cid, d))

    # ── Print report ──
    print('\n' + '=' * 80)
    print('Phase B audit — vault ↔ ERP cid 绑定健康度')
    print('=' * 80)

    print(f'\n✓ Section A — Clean ({len(sec_A_clean)})')
    print(f'    vault cid 在 ERP 且 客户姓名 与 vault folder 一致')

    if sec_B_name_mismatch:
        print(f'\n⚠ Section B — Name mismatch ({len(sec_B_name_mismatch)}) — 需要核对')
        for vs, cid, erp_names in sec_B_name_mismatch[:30]:
            print(f'  • vault: {vs["name"]:<12}  cid={cid}  ERP 客户姓名={erp_names}')
        if len(sec_B_name_mismatch) > 30:
            print(f'  ...等 {len(sec_B_name_mismatch) - 30} 个')
    else:
        print(f'\n✓ Section B — Name mismatch: 0 ✓')

    if sec_C_cid_not_in_erp:
        print(f'\n⚠ Section C — vault cid 在 ERP 找不到 ({len(sec_C_cid_not_in_erp)})')
        for vs, cid in sec_C_cid_not_in_erp[:20]:
            print(f'  • {vs["name"]:<12} cid={cid}')
        if len(sec_C_cid_not_in_erp) > 20:
            print(f'  ...等 {len(sec_C_cid_not_in_erp) - 20} 个')
    else:
        print(f'\n✓ Section C — vault cid 在 ERP 找不到: 0 ✓')

    if sec_D_no_cid:
        print(f'\n⚠ Section D — vault 学生没绑 cid (非私单) ({len(sec_D_no_cid)})')
        by_mid = defaultdict(list)
        for vs in sec_D_no_cid:
            by_mid[vs["mid"] or '(无)'].append(vs)
        for mid, lst in sorted(by_mid.items(), key=lambda x: -len(x[1])):
            print(f'  -- 中期={mid} ({len(lst)}) --')
            for vs in lst[:5]:
                print(f'    · {vs["name"]:<12}  intake={vs["intake"]:<12}  stage={vs["stage"]}')
            if len(lst) > 5:
                print(f'    ...等 {len(lst) - 5}')
    else:
        print(f'\n✓ Section D — vault 学生没绑 cid: 0 ✓')

    if sec_E_multi_cid:
        print(f'\nℹ Section E — vault folder 绑多个 cid ({len(sec_E_multi_cid)}) — 通常是 sibling/alias，确认是合理')
        for vs in sec_E_multi_cid[:15]:
            print(f'  • {vs["name"]:<14} ({len(vs["cids"])} cids: {vs["cids"]})')
        if len(sec_E_multi_cid) > 15:
            print(f'  ...等 {len(sec_E_multi_cid) - 15} 个')

    # ── Section F ──
    print(f'\n❌ Section F — ERP cid 没绑到 vault ({len(sec_F_orphan)})')
    # Categorize
    f_active_app_inhouse = []   # 26F+ 留学+在带, 中期在 inhouse
    f_active_app_other = []     # 26F+ 留学+在带, 中期不在 inhouse (or 空)
    f_old = []                  # 历史 / 已退费 / 非留学
    for cid, d in sec_F_orphan:
        try:
            y = int(d['app_intake'] or d['intake'] or 0)
        except (TypeError, ValueError):
            y = 0
        if y >= 2026 and d['has_active'] and d['has_app']:
            inh = bool(d['app_mid_set'] & INHOUSE_ADVISORS)
            (f_active_app_inhouse if inh else f_active_app_other).append((cid, d))
        else:
            f_old.append((cid, d))

    print(f'  🔴 26F+ 留学+在带, 中期 in-house顾问 ({len(f_active_app_inhouse)}) — 待 onboard:')
    by_mid = defaultdict(list)
    for cid, d in f_active_app_inhouse:
        primary_mid = next(iter(d['app_mid_set'] & INHOUSE_ADVISORS), None)
        by_mid[primary_mid].append((cid, d))
    for mid, lst in sorted(by_mid.items(), key=lambda x: -len(x[1])):
        print(f'    -- 中期={mid} ({len(lst)}) --')
        for cid, d in lst[:8]:
            nm = next(iter(d['names']))
            print(f'      · {nm:<10} intake={d["app_intake"] or d["intake"]}F cid={cid}')
        if len(lst) > 8:
            print(f'      ...等 {len(lst) - 8}')

    print(f'\n  🟡 26F+ 留学+在带, 中期 ≠ in-house (或空) ({len(f_active_app_other)}) — 多半外部 / 没分配:')
    by_mid = defaultdict(int)
    for cid, d in f_active_app_other:
        for m in (d['app_mid_set'] or {'(无)'}):
            by_mid[m] += 1
    for mid, n in sorted(by_mid.items(), key=lambda x: -x[1])[:10]:
        print(f'    · {mid}: {n}')
    print(f'  · 历史/已结案/退费/非留学 ({len(f_old)}) — 通常可忽略')

    # ── Summary ──
    print('\n' + '=' * 80)
    print('SUMMARY')
    print(f'  ✓ A. Clean:                       {len(sec_A_clean)}')
    print(f'  ⚠ B. Name mismatch:               {len(sec_B_name_mismatch)}')
    print(f'  ⚠ C. cid 不在 ERP:                {len(sec_C_cid_not_in_erp)}')
    print(f'  ⚠ D. vault 无 cid (非私单):       {len(sec_D_no_cid)}')
    print(f'  ℹ E. vault 多 cid (合理):         {len(sec_E_multi_cid)}')
    print(f'  ❌ F. ERP orphan cid:             {len(sec_F_orphan)}')
    print(f'      └ 🔴 in-house 应 onboard:     {len(f_active_app_inhouse)}')
    print(f'      └ 🟡 外部/无中期:              {len(f_active_app_other)}')
    print(f'      └ · 历史/退费 (忽略):          {len(f_old)}')
    print('=' * 80)

    # Compute切换风险
    print('\n📋 切到 cid-driven 前的 BLOCKER:')
    if sec_B_name_mismatch:
        print(f'   ⚠ {len(sec_B_name_mismatch)} 个 name mismatch — 需先核对是 alias 还是误绑')
    if sec_C_cid_not_in_erp:
        print(f'   ⚠ {len(sec_C_cid_not_in_erp)} 个 vault cid 在 ERP 找不到 — 可能错绑，要解')
    if sec_D_no_cid:
        print(f'   ⚠ {len(sec_D_no_cid)} 个 vault 学生没 cid — 切到 cid-driven 后会跟 ERP 失联')
    if f_active_app_inhouse:
        print(f'   🔴 {len(f_active_app_inhouse)} 个 in-house 应 onboard — 不切换也建议处理')
    if not (sec_B_name_mismatch or sec_C_cid_not_in_erp):
        print('   ✓ 无 blocker 阻止 Phase B 切换')


if __name__ == '__main__':
    main()
