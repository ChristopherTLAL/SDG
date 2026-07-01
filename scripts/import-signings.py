#!/usr/bin/env python3
"""
import-signings.py — single source of truth for contract data.

Source (default, 2026-06-30 onward): the newest SINGLE .xlsx with a 签约明细 sheet in
~/Downloads (+ _工作运营/). Workflow now exports one full 2023-06→今 book each time —
全量自带每条合同的最新状态（退费/终止/完成），取最新那份即可。Pass an explicit path to
force a specific file; pass --union for the OLD multi-file-union behavior (union every
signing xlsx by 合同编号, newest file wins) as a fallback.

Three things it does, all idempotent:

1. **Update existing vault students' YAML**:
   - add `客户ID: [...]` (array, supports sibling-combined folders)
   - add `合同签约人: [...]` (alias for special cases like Kimi+Byran → LI KIMI)
   - merge `合同明细` from signing data (DEFAULT; `--rewrite` replaces instead,
     dropping vault contracts absent from all files; legacy 预收余额/已收/已交付 dropped)
   - new contract entry shape: 名称 / 大类 / 签约日期 / 签约金额 / 合同状态 / 合同编号

2. **Onboard new vault folders** for advisors who lack vault coverage
   (see ONBOARD_ADVISORS list — currently 钟婷婷 / 古淑婷 / 高幸玲 / 王姝琰).
   - skips if folder already exists
   - if name collides with existing vault folder → suffix with " (中期顾问)"
   - intake-year-based 当前进度 (2025F or earlier = 已结案, 2026F = 后期在途, 2027F+ = 中期在途)

3. **Push signing data to Supabase contracts table** (UPSERT by 合同编号), NARROWED
   to「在库学生的全部合同」+「全公司非留学产品合同 (is_application=false)」. The latter
   feeds the kanban 月度产品节奏 widget (company-wide, 2026-05-09 directive). 别部门的
   【留学】合同不入库 (nothing reads them). 此口径与 kanban widget + purge SQL 三者一致。

After running this, also rerun:
   /usr/local/bin/python3 scripts/import-erp-comms.py
   (so newly-onboarded students get their ERP historical 沟通 archived)
   node scripts/sync-students-to-supabase.mjs
   (so dashboard sees the changes)
"""
import json
import os
import re
import sys
import time
import zipfile
import urllib.request
import urllib.error
from collections import defaultdict, Counter
from datetime import datetime, date, timedelta
from pathlib import Path

import openpyxl


# ── Config ───────────────────────────────────────────────────────────────

VAULT = Path('/Users/shijie/Obsidian/规划看板/01_Student')
SDG_HTML = Path('/Users/shijie/Code/sdg-html')
ENV_FILE = SDG_HTML / '.env'

# Resolve signing xlsx source(s):
#   1. CLI path arg → use EXACTLY that one file (explicit single-file override).
#   2. Default (no arg) → MULTI-FILE UNION: find every .xlsx that contains a 签约明细
#      sheet (detection by SCHEMA, not filename — so a weekly export named anything,
#      e.g. 53周.xlsx / 本周签约.xlsx, is picked up automatically) across ~/Downloads/
#      + ~/Downloads/_工作运营/, union contracts by 合同编号, newest file wins on
#      conflict. A weekly slice can never drop contracts that live only in the
#      full-history export; an old backfill just adds its missing contracts.
#      "几个文档都要看" — all of them, every run, regardless of filename.
def _has_signing_sheet(path):
    """Cheap, filename-agnostic check: does this .xlsx contain a 签约明细 sheet?
    Reads only xl/workbook.xml from the zip (sheet names live there) — no full
    workbook parse — so scanning a Downloads folder of 100+ xlsx stays sub-second."""
    try:
        with zipfile.ZipFile(path) as z:
            return '签约明细' in z.read('xl/workbook.xml').decode('utf-8', 'ignore')
    except Exception:
        return False

def _all_signing_xlsx():
    search_dirs = [
        Path.home() / 'Downloads',
        Path.home() / 'Downloads' / '_工作运营',
    ]
    cands = []
    for d in search_dirs:
        if not d.exists():
            continue
        for p in d.glob('*.xlsx'):
            if not p.name.startswith('~$') and _has_signing_sheet(p):
                cands.append(p)
    if not cands:
        raise FileNotFoundError(
            'No .xlsx with a 签约明细 sheet found in ~/Downloads or ~/Downloads/_工作运营')
    # oldest → newest by mtime, so the newest file wins on union.
    return sorted(cands, key=lambda p: p.stat().st_mtime)

def _newest_signing_xlsx():
    """Default source (2026-06-30 onward): the SINGLE newest .xlsx with a 签约明细
    sheet. Workflow changed to「每次导出一份 2023-06→今 的全量」——全量自带每条合同
    的最新状态（退费/终止/完成），取最新那一份即可，不必再 union 一堆旧切片（旧并集
    会把恰好躺在 ~/Downloads 的历史大表也卷进来）。用 --union 恢复旧的多文件并集行为。"""
    cands = []
    for d in (Path.home() / 'Downloads', Path.home() / 'Downloads' / '_工作运营'):
        if not d.exists():
            continue
        for p in d.glob('*.xlsx'):
            if not p.name.startswith('~$') and _has_signing_sheet(p):
                cands.append(p)
    if not cands:
        raise FileNotFoundError(
            'No .xlsx with a 签约明细 sheet found in ~/Downloads or ~/Downloads/_工作运营')
    return [max(cands, key=lambda p: p.stat().st_mtime)]

# Flags:
#   --dry-run    : read everything, compute changes, print summary, NO writes.
#   --rewrite    : AUTHORITATIVE — REPLACE each vault 合同明细 from the file union,
#                  DROPPING any vault contract absent from every file. Use only when
#                  you deliberately want to prune stale/cancelled contracts.
#   (default)    : MERGE the file union into existing vault 合同明细 by 合同编号
#                  (file wins on conflict, vault-only contracts are KEPT). This is
#                  the safe weekly behavior — it never loses a vault 档案 contract.
#   --incremental: accepted for back-compat but now a NO-OP (merge IS the default).
#   --union      : OLD default — scan ~/Downloads + _工作运营 and union every signing
#                  xlsx by 合同编号. Now opt-in only; default = newest single full-export.
DRY_RUN     = '--dry-run' in sys.argv
REWRITE     = '--rewrite' in sys.argv
UNION       = '--union' in sys.argv
INCREMENTAL = not REWRITE          # merge-into-vault default; --rewrite opts out
_FLAGS = ('--dry-run', '--incremental', '--rewrite', '--union')
_pos_args = [a for a in sys.argv[1:] if a not in _FLAGS]

EXPLICIT_FILE = None               # set when a CLI path arg forces single-file mode
if _pos_args and _pos_args[0] not in ('-h', '--help'):
    EXPLICIT_FILE = Path(_pos_args[0]).expanduser()
    if not EXPLICIT_FILE.exists():
        print(f'ERROR: {EXPLICIT_FILE} not found', file=sys.stderr)
        sys.exit(1)
elif _pos_args:
    print(__doc__)
    print(f'\nUsage: {sys.argv[0]} [--dry-run] [--rewrite] [--union] [path/to/full-export.xlsx]')
    print(f'\nNo arg  → newest SINGLE .xlsx with a 签约明细 sheet in ~/Downloads (+ _工作运营/).')
    print(f'          Workflow: 每次导出一份 2023-06→今 全量；全量自带最新状态，取最新那份。')
    print(f'path    → use EXACTLY that one file (single-file override).')
    print(f'--union → OLD behavior: union EVERY signing xlsx in ~/Downloads by 合同编号')
    print(f'          (filename-agnostic; newest file wins). Opt-in fallback only.')
    print(f'--dry-run → preview changes without writing to vault or Supabase.')
    print(f'--rewrite → authoritative: REPLACE vault 合同明细 from files, dropping vault')
    print(f'            contracts absent from all files. Default = merge (never drops).')
    sys.exit(0)


# Aliases — vault folder name → list of contract signer 客户姓名
# Add new entries here as you discover them.
ALIASES = {
    'Kimi+Byran':  ['LI KIMI', 'Bryan Wang'],
    'Sophia+Amy':  ['LI XINYING', 'Sophia Wang'],
    'Betty':       ['陈碧涵'],
}

# 重名 disambiguation — vault folder name → exact 客户ID to use
# (only one ID; for sibling folders use ALIASES instead).
DISAMBIG = {
    '李子萱':           'P24AABbujT1xmUc95246',
    '周鑫':             'P24AABXx3yxxmUc90018',
    '胡斌':             'P24AABpPeLNxmUc92018',
    '李想 (钟婷婷)':    'P24AAB2YSCFxmUc95258',  # vault folder 后缀，避免每次跑都尝试 re-onboard
    '王思涵':           'P24AAB73Y11xmUc94528',  # vault 实际只有 27F (古淑婷线)，另一个 23F cid 是误绑
    '杨阳':             'P24AABA7789xmUc94681',  # 26F 古淑婷线；另一 c90535=23 已终止别人 (260611 dedup)
    '李家乐':           'P24AABeGnzhxmUc90051',  # 26F 钟婷婷线；另一 c90415=22 别人 (260611 dedup)
    '张凯':             'P24AABQWuHFxmUc93532',  # 27F 古淑婷线；另一 c93818=25 葛倩别人 (260611 dedup)
    '周佳怡':           'P24AABRpaUdxmUc93521',  # 24F 钟婷婷线；另一 c92526=27F 张文心新人,待前期确认 (260611 dedup)
    '彭涛':             'P24AABzWrV9xmUc90018',  # 27F 现役；另一 c90115=21 加拿大已完成别人 (260611)
    '周晨阳':           'P24AABYdP0ZxmUc90013',  # 25F；另一 c9225X=21 英国别人 (260611)
}

# VIP (private-contract; 旧称「私单」) detection is YAML-driven (2026-05-17 onward).
# vault .md frontmatter `合同` 字段含 `VIP` / `VIP（非公司合同）`（或 legacy `私单`）的学生
# 被识别为 VIP 并跳过 ERP 匹配 / 不被 import-signings.py 改写 YAML。
# 不要再 hardcode 名字 list — 在 vault YAML 改 `合同: [VIP]` 即可。
# 同时保留 legacy 私单 值以兼容尚未迁移的数据。
PRIVATE_CONTRACT_LABELS = ('VIP', 'VIP（非公司合同）', '私单', '私单（非公司合同）')

# Onboard new vault folders for these advisors' uncovered students:
ONBOARD_ADVISORS = ['钟婷婷', '古淑婷', '高幸玲', '王姝琰']

# Auto-onboard filter — only build folder if cid is still in-scope:
#   - intake_year ≥ ONBOARD_MIN_INTAKE (skip already-closed cohorts)
#   - at least one contract is NOT 已终止/退费 (skip refund-only cids)
# Pre-filter saves us from /tmp/onboard_X.py one-shot scripts to avoid脏 onboards.
# Bump ONBOARD_MIN_INTAKE每年 (e.g. 2027 in fall 2026).
ONBOARD_MIN_INTAKE = 2026

TODAY = datetime.now().strftime('%Y-%m-%d')


# ── Helpers ──────────────────────────────────────────────────────────────

def cleanup_addon_name(s):
    """Clean a 合同模板名称 → stable 大类. Strips contract-instance / boilerplate
    cruft while PRESERVING meaningful content (price/tier parens, sub-plan letters).

    Rewritten 2026-07-01 (was: leaked cruft into 大类 → 681+ near-dup categories,
    e.g. '美高Summer School合同-OA4907032', '-1对1精英定制', '财年韩国留学…'). Verified
    against all 1691 distinct template_names: 0 residual cruft, 0 regressions,
    specific-rule families (跃领/格物/菁英/启航…) + price/tier parens untouched.

    Removes:
      - leading time qualifiers: 2024 / 23财 / 22财年 / 24暑假 / 2024-2025学年 / 24-25学年
      - OA / order codes anywhere: US-OA123 / 合同-OA123 / 合同OA123#1 / …-OA…—2D / …#N-M
      - contract-type boilerplate: 项目服务合同 / 服务合同 / 服务协议 / 项目服务 / 合同 / 项目
        (bare 合同 preceded by an ASCII letter is KEPT — sub-plan labels A合同/B合同)
    Then collapses stray double/em dashes and trims leftover separators.
    """
    s = s.strip()
    # 1. leading time qualifiers
    s = re.sub(r'^\d{4}\s*[-—]\s*\d{4}\s*学年\s*', '', s)   # 2024-2025学年 (school-year range)
    s = re.sub(r'^\d{2}\s*[-—]\s*\d{2}\s*学年\s*', '', s)   # 24-25学年
    s = re.sub(r'^\d{4}\s*财?\s*年?\s*', '', s)             # 2024 / 2024财 / 2024财年
    s = re.sub(r'^\d{2}\s*财?\s*年?\s*', '', s)             # 23财 / 22财年 / 24暑假 (all leading 2-digit = year)
    s = re.sub(r'^财\s*年?\s*', '', s)                      # stray 财年 leftover (e.g. 23财财年…)
    # 2. OA / order codes (token = optional [A-Z]{2}- or 合同(-) prefix + OA + digits
    #    + optional #n / -n / —n / -<short alnum> instance suffixes)
    OA = r'(?:[A-Z]{2}-|合同[-]?)?OA\s*\d+(?:\s*[#\-—]\s*[0-9A-Za-z]{1,4})*'
    s = re.sub(r'\s*' + OA + r'\s*$', '', s)                         # a) trailing code (eats leading ws)
    s = re.sub(r'(?:合同\s*)?[-—]?\s*(?:[A-Z]{2}-)OA\s*\d+', '', s)   # b) mid-string coded (US-OA…)
    s = re.sub(r'合同\s*OA\s*\d+(?:[#\-—][0-9A-Za-z]{1,4})*', '', s)  #    mid-string 合同OA…
    # 3. contract-type boilerplate
    s = re.sub(r'(项目服务合同|项目服务协议|项目服务|服务协议|服务合同|项目合同)', '', s)  # compounds anywhere
    s = re.sub(r'(?<![A-Za-z])合同(?=\s*[（(\-—]|\s*$)', '', s)   # bare 合同 (protect A合同/B合同)
    s = re.sub(r'项目(?=\s*[（(\-—]|\s*$)', '', s)                # bare 项目 (always boilerplate)
    # 4. tidy
    s = re.sub(r'\s*[-—]\s*[-—]\s*', '-', s)   # collapse -- / —— left mid-string
    s = re.sub(r'\s+', ' ', s)
    s = s.strip(' \t-—·、,，:：')
    return s.strip()


def category(s):
    """Map 合同模板名称 → 大类. Same logic as sync-financial-to-vault.py."""
    if not s:
        return 'unknown'
    if "跃领" in s and "博士" in s: return "跃领-博士版"
    if "跃领" in s and "本科" in s: return "跃领-本科版"
    if "跃领" in s and "科研" in s: return "跃领-科研addon"
    if "跃领" in s: return "跃领"
    if "格物" in s and ("半年" in s or "（半年）" in s): return "格物-半年"
    if "格物" in s and ("一年" in s or "（一年）" in s): return "格物-一年"
    if "格物" in s: return "格物-一年"
    if "精英预备" in s: return "精英预备课程"
    if "菁英" in s and "博士" in s: return "博士"
    if "菁英" in s: return "菁英"
    if "亚洲英文授课博士" in s or "亚洲英文授课硕士" in s: return "亚洲博士"
    if "亚洲英语系高端" in s or "欧亚" in s or "亚英高端" in s: return "亚洲英语系高端"
    if "亚洲英文授课" in s and ("研究生" in s or "本科" in s or "授课式" in s): return "亚洲英语系高端"
    if "新港澳" in s or "新港" in s or "中国港澳" in s: return "新港澳联申"
    if "美国研究生启航" in s: return "启航"
    if "美国研究生快捷当季" in s or "美国研究生快捷" in s: return "美国研究生快捷"
    if "美研尊享" in s or "美国研究生尊享" in s: return "尊享"
    if ("澳大利亚" in s or "加拿大" in s) and "签证" in s: return cleanup_addon_name(s)
    if "澳大利亚" in s or "澳新" in s: return "澳洲申请"
    if "加拿大" in s: return "加拿大申请"
    if "云中学" in s: return "云中学"
    if "香港" in s or "港-" in s or "港A" in s or "港授课" in s: return "港申请"
    if "英国本科" in s: return "英国本科"
    if "英国研究生" in s and "博士" in s: return "英国硕博"
    if "美国线上中学" in s: return "美国本科预备"
    if "欧洲英语系博士" in s: return "欧洲博士"
    return cleanup_addon_name(s)


def is_yes(v):
    return v == '是' or v is True or v == 1 or v == '1'


def parse_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    if not s:
        return None
    return s[:10]


def progress_for_intake(intake_year):
    """Map intake_year to 当前进度. As of today (2026):
        ≤ 2025: 已结案 (graduated / enrolled)
        2026:   后期在途 (final stretch)
        ≥ 2027: 中期在途 (still in mid stage)
    """
    if intake_year is None:
        return '中期在途'
    try:
        y = int(intake_year)
    except (TypeError, ValueError):
        return '中期在途'
    if y <= 2025: return '已结案'
    if y == 2026: return '后期在途'
    return '中期在途'


def safe_yaml_str(s):
    """YAML-safe string (quote if contains special chars)."""
    if s is None:
        return ''
    s = str(s)
    if any(c in s for c in [':', '#', '@', '`', '"', "'", '[', ']', '{', '}', '\n']):
        return '"' + s.replace('"', '\\"') + '"'
    return s


# ── Read .env for Supabase ──────────────────────────────────────────────

def load_env():
    env = {}
    if not ENV_FILE.exists():
        print(f'⚠ .env not found at {ENV_FILE}; Supabase upload will be skipped', file=sys.stderr)
        return env
    for line in ENV_FILE.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith('#') or '=' not in line:
            continue
        k, v = line.split('=', 1)
        env[k.strip()] = v.strip().strip('"').strip("'")
    return env


# ── Read signing xlsx ───────────────────────────────────────────────────

def read_signings():
    """Read contract rows. Default = union of ALL 客户签约明细 exports (newest file
    wins per 合同编号); a CLI path forces a single file. Rows are normalized to a
    canonical column order (by header name) so exports with slightly different
    column layouts still union correctly. Keeps rows with 客户姓名+客户ID+合同编号."""
    # 新版「Sign_Archiving_Details_dept」自助导出表的列名 → 旧版 canonical 列名。
    # 仅当该文件里 canonical 名不存在时才改名（避免与旧大表已有列冲突）；
    # 旧表列名不在 map → 原样通过，所以新旧文件能无缝 union。
    HEADER_ALIAS = {
        '合同号': '合同编号', '日期': '签约时间', '签约金额': '签约金额（实时）',
        '合同模板': '合同模板名称', '留学申请': '是否包含留学申请',
        '语言培训': '是否包含语言培训', '申请入学年': '入学年',
    }
    if EXPLICIT_FILE:
        files = [EXPLICIT_FILE]
        label = 'single file (explicit path)'
    elif UNION:
        files = _all_signing_xlsx()
        label = f'{len(files)}-file union (--union)'
    else:
        files = _newest_signing_xlsx()
        label = 'newest single full-export'
    print(f'📖 Reading {label}: ' + ', '.join(f.name for f in files))
    canonical = []          # header names, first-seen order
    COL = {}                # header name → canonical index
    by_contract = {}        # 合同编号 → row (normalized to canonical order)
    NEED = ('客户姓名', '客户ID', '合同编号')
    for f in files:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb['签约明细']
        try:
            ws.reset_dimensions()   # 部门自助导出表 dimension 元信息常坏 → read_only 只读到首列；强制按真实单元格重扫
        except Exception:
            pass
        fheader_raw = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        _present = {str(h).strip() for h in fheader_raw if h}
        fheader = [
            HEADER_ALIAS[str(h).strip()]
            if (h and str(h).strip() in HEADER_ALIAS and HEADER_ALIAS[str(h).strip()] not in _present)
            else h
            for h in fheader_raw
        ]
        fcol = {h: i for i, h in enumerate(fheader) if h}
        for h in fheader:                       # extend canonical with new columns
            if h and h not in COL:
                COL[h] = len(canonical); canonical.append(h)
        if not all(k in fcol for k in NEED):
            print(f'   ⚠ skip {f.name}: missing {[k for k in NEED if k not in fcol]}')
            wb.close(); continue
        ci, idi, ni = fcol['客户姓名'], fcol['客户ID'], fcol['合同编号']
        n = 0
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not (r and len(r) > ni and r[ci] and r[idi] and r[ni]):
                continue
            norm = [None] * len(canonical)      # remap to canonical order by name
            for h, i in fcol.items():
                if i < len(r):
                    norm[COL[h]] = r[i]
            by_contract[r[ni]] = norm           # newest file wins (oldest→newest)
            n += 1
        wb.close()
        print(f'   · {f.name}: {n} rows')
    # 全量哨兵：默认工作流期望一份 2023-06→今 的全量导出。日期跨度过短 = 多半误传了
    # 月度切片；merge 模式不会丢数据，但旧合同状态不会被刷新，故提醒一声。
    _sidx = COL.get('签约时间')
    if _sidx is not None:
        _ds = [str(r[_sidx])[:10] for r in by_contract.values()
               if _sidx < len(r) and r[_sidx]]
        if _ds:
            _lo, _hi = min(_ds), max(_ds)
            print(f'   📅 签约时间跨度: {_lo} → {_hi}')
            try:
                if (date.fromisoformat(_hi) - date.fromisoformat(_lo)).days < 365:
                    print('   ⚠ 跨度 < 1 年 —— 看起来像月度切片而非 2023-06 起的全量导出；'
                          '确认是否误传（merge 不丢数据，但旧合同状态不刷新）。')
            except Exception:
                pass
    width = len(canonical)
    rows = [tuple(row) + (None,) * (width - len(row)) for row in by_contract.values()]
    print(f'   = {len(rows)} unique contracts (unioned by 合同编号 across {len(files)} file[s])')
    return rows, COL


# ── Read existing vault ─────────────────────────────────────────────────

def _yaml_scalar(yaml_text, key):
    """读 YAML frontmatter 里 `key: value` 的标量值. 空值返回 ''.
    用 `[ \\t]*` 不用 `\\s*` 以避免吃下一行的换行 + 字段名."""
    m = re.search(rf'^{re.escape(key)}:[ \t]*(.*?)$', yaml_text, re.MULTILINE)
    if not m:
        return None
    return m.group(1).strip()


def _yaml_is_private(text):
    """True iff frontmatter `合同` 字段的某个 list 条目 *精确等于* PRIVATE_CONTRACT_LABELS
    中的任一标签. 驱动私单识别 — 取代之前 hardcode 的 PRIVATE_NO_SIGNING name list。

    必须 per-item 精确匹配, 不能拿整串做 substring: 否则像
    「新东方国际竞赛学术指导VIP小班（面授落地）」这种【产品名】里含 "VIP" 子串的普通合同
    会被误判成私单 (2026-06-22 茅晨曦 bug: 被误跳过 ERP 匹配 → cid 不绑 → onboard 又
    重建 "茅晨曦 (王姝琰)" 空壳重复档)."""
    m = re.match(r'^---\s*\n(.*?)\n---', text, re.DOTALL)
    if not m:
        return False
    val = _yaml_scalar(m.group(1), '合同')
    if not val:
        return False
    val = val.strip()
    if val.startswith('[') and val.endswith(']'):
        items = [s.strip().strip('"').strip("'") for s in val[1:-1].split(',')]
    else:
        items = [val.strip('"').strip("'")]
    return any(item in PRIVATE_CONTRACT_LABELS for item in items)


def _yaml_advisor_fields(text):
    """从 vault .md frontmatter 抽取 中期/前期/后期 顾问 (primary, 多顾问取第一个).
    返回 dict {'mid': ..., 'early': ..., 'late': ...}, 找不到的 key 缺失.
    用于 Phase A: contracts 表 mid/early/late_advisor 用 vault 值覆盖 ERP 值."""
    m = re.match(r'^---\s*\n(.*?)\n---', text, re.DOTALL)
    if not m:
        return {}
    yaml = m.group(1)
    out = {}
    for yaml_key, out_key in [('中期顾问', 'mid'), ('前期顾问', 'early'), ('后期顾问', 'late')]:
        val = _yaml_scalar(yaml, yaml_key)
        if not val:
            continue
        if val.startswith('['):
            items = [s.strip().strip('"').strip("'") for s in val.strip('[]').split(',')]
            items = [i for i in items if i]
            if items:
                out[out_key] = items[0]
        else:
            out[out_key] = val.strip('"').strip("'")
    return out


def read_vault_students():
    students = []
    for p in sorted(VAULT.iterdir()):
        if not p.is_dir() or p.name.startswith('_'):
            continue
        md = p / f'{p.name}.md'
        if not md.exists():
            continue
        text = md.read_text(encoding='utf-8', errors='ignore')
        students.append({
            'folder': p,
            'name': p.name,
            'md_path': md,
            'is_private': _yaml_is_private(text),
            'advisors': _yaml_advisor_fields(text),  # Phase A: cache 顾问字段，避免重读
        })
    return students


# ── Build customer_id → student folder mapping ─────────────────────────

def build_customer_to_folder_map(rows, COL, vault_students):
    """Returns dict[customer_id] = folder_name (or None if unmapped)."""
    name_to_ids = defaultdict(set)
    for r in rows:
        name_to_ids[r[COL['客户姓名']]].add(r[COL['客户ID']])

    cid_to_folder = {}
    folder_to_cids = defaultdict(set)
    folder_signers = defaultdict(set)  # folder → set of 客户姓名 used in signing

    private_folders = sorted(vs['name'] for vs in vault_students if vs['is_private'])
    if private_folders:
        print(f'🔒 私单 folders 跳过 ERP 匹配 ({len(private_folders)}): {private_folders}')

    # 重名保护: ERP 中 客户姓名 X → >1 个不同 cid, 且 vault 有同名非私单 folder,
    # 但没有 DISAMBIG entry 显式选 cid → 不自动绑定（防止把两个真实不同客户的
    # 合同合并写进同一个 vault 文件夹）。这种情况要么加 DISAMBIG，要么走 onboard
    # 自动建 "X (中期顾问名)" suffix folder。
    aliased_signers = {s for signers in ALIASES.values() for s in signers}
    ambiguous_names = set()
    for name, ids in name_to_ids.items():
        if len(ids) <= 1:
            continue
        if name in DISAMBIG or name in aliased_signers:
            continue
        vault_match = [vs for vs in vault_students
                       if vs['name'] == name and not vs['is_private']]
        if vault_match:
            ambiguous_names.add(name)
            print(f'⚠ 重名: ERP 有 {len(ids)} 个 cid 客户姓名 == "{name}"  '
                  f'→ vault folder 不自动绑定。要绑定请加 DISAMBIG: {sorted(ids)}')

    for vs in vault_students:
        name = vs['name']

        if vs['is_private']:
            continue  # 私单 — 跟 ERP 完全隔离

        # ALIASES first (sibling-combined / English-name signed)
        if name in ALIASES:
            for signer in ALIASES[name]:
                for cid in name_to_ids.get(signer, set()):
                    cid_to_folder[cid] = name
                    folder_to_cids[name].add(cid)
                    folder_signers[name].add(signer)
            continue

        # DISAMBIG (重名 with explicit ID choice)
        if name in DISAMBIG:
            cid = DISAMBIG[name]
            cid_to_folder[cid] = name
            folder_to_cids[name].add(cid)
            folder_signers[name].add(name)
            continue

        if name in ambiguous_names:
            continue  # 重名未消歧 — 跳过自动绑定

        # Normal direct name match
        ids = name_to_ids.get(name, set())
        for cid in ids:
            cid_to_folder[cid] = name
            folder_to_cids[name].add(cid)
            folder_signers[name].add(name)

    return cid_to_folder, folder_to_cids, folder_signers


# ── Identify new students to onboard ────────────────────────────────────

def identify_onboard_candidates(rows, COL, cid_to_folder, vault_folder_names):
    """For each ONBOARD_ADVISORS, find customers (留学申请=是) not covered by vault.

    Pre-filter (since 2026-05-17):
    - intake_year ≥ ONBOARD_MIN_INTAKE (skip closed cohorts e.g. 25F 已结案)
    - has_active = 至少一条合同非 终止/退费 (skip 全退费 cids)
    Without these, the script would silently onboard历史死案，produces stale
    vault folders that need to be hand-cleaned. 这个过滤等价于之前 /tmp/onboard_X.py
    的"26F+ active 留学" 筛选规则，现在内置。
    """
    # Track per-cid: latest intake + has_active across all this cid's contracts
    per_cid = {}  # cid → {first_advisor_row, intake_pref, has_active, names, count, total}
    for r in rows:
        adv = r[COL['中期顾问']]
        if adv not in ONBOARD_ADVISORS:
            continue
        if not is_yes(r[COL['是否包含留学申请']]):
            continue
        cid = r[COL['客户ID']]
        if cid in cid_to_folder:
            continue  # already covered
        nm = r[COL['客户姓名']]
        intake = r[COL['入学年']]
        status = str(r[COL['合同状态']] or '')
        is_active = bool(status) and '终止' not in status and '退' not in status

        d = per_cid.setdefault(cid, {
            'cid': cid, 'name': nm, 'advisor': adv,
            'intake': intake, 'has_active': False,
            'count': 0, 'total': 0.0,
        })
        d['count'] += 1
        d['total'] += r[COL['签约金额（实时）']] or 0
        if intake and (not d['intake']):
            d['intake'] = intake
        d['has_active'] = d['has_active'] or is_active

    # Apply onboard filter — drop 已结案 cohorts + all-refunded cids
    candidates = {}
    skipped_old, skipped_dead = 0, 0
    for cid, c in per_cid.items():
        try:
            y = int(c['intake'])
        except (TypeError, ValueError):
            y = 0
        if y < ONBOARD_MIN_INTAKE:
            skipped_old += 1
            continue
        if not c['has_active']:
            skipped_dead += 1
            continue
        candidates[cid] = c
    if skipped_old or skipped_dead:
        print(f'   📥 onboard filter: dropped {skipped_old} (<{ONBOARD_MIN_INTAKE}F) + {skipped_dead} (全退费)')

    # Detect collisions with existing vault folders + within candidate list
    final = []
    for cid, c in candidates.items():
        target_folder_name = c['name']
        # If collides with existing vault → suffix with " (advisor)"
        if target_folder_name in vault_folder_names:
            target_folder_name = f"{c['name']} ({c['advisor']})"
        c['target_folder'] = target_folder_name
        final.append(c)
    return final


# ── YAML rewrite: 合同明细 ──────────────────────────────────────────────

def read_existing_contracts_from_yaml(md_path):
    """Parse existing 合同明细 entries from vault md frontmatter.

    Returns list of dicts with the same shape as `contracts` rows we build:
      template_name / category / signed_at / signed_amount / status / contract_id

    Returns [] if file doesn't exist or has no 合同明细 entries. Tolerant of
    missing/messy YAML — anything unparseable is skipped silently so an existing
    bad entry doesn't block a new write.
    """
    if not md_path.exists():
        return []
    text = md_path.read_text(encoding='utf-8', errors='ignore')
    fm = re.match(r'(?s)^---\s*\n(.*?)\n---', text)
    if not fm:
        return []
    fm_body = fm.group(1)
    # Find the 合同明细 block — everything from `合同明细:` until the next
    # top-level YAML key (a line starting with non-whitespace + `:`).
    m = re.search(
        r'^合同明细:\s*\n((?:[ \t]+.*\n?)*)',
        fm_body,
        re.MULTILINE,
    )
    if not m:
        return []
    block = m.group(1)
    # Parse list-of-dict entries. Each entry begins with `  - 名称: ...`.
    entries = []
    cur = None
    for line in block.splitlines():
        if not line.strip():
            continue
        list_item = re.match(r'^\s*-\s*名称:\s*(.*?)\s*$', line)
        kv = re.match(r'^\s+(\S[^:]*?):\s*(.*?)\s*$', line)
        if list_item:
            if cur:
                entries.append(cur)
            cur = {'template_name': list_item.group(1).strip('"').strip("'")}
        elif kv and cur is not None:
            key, val = kv.group(1), kv.group(2).strip('"').strip("'")
            if key == '大类':
                cur['category'] = val
            elif key == '签约日期':
                cur['signed_at'] = val
            elif key == '签约金额':
                try:
                    cur['signed_amount'] = float(val) if '.' in val else int(val)
                except ValueError:
                    cur['signed_amount'] = val
            elif key == '合同状态':
                cur['status'] = val
            elif key == '合同编号':
                cur['contract_id'] = val
    if cur:
        entries.append(cur)
    # Filter out entries without 合同编号 (can't dedupe them safely)
    return [e for e in entries if e.get('contract_id')]


def merge_contracts_by_id(existing, new):
    """Merge two contract lists, file (new) wins on 合同编号 conflict.
    Order: existing first (preserves vault order), then new entries
    whose 合同编号 wasn't in existing.
    """
    existing_ids = {c['contract_id']: i for i, c in enumerate(existing)}
    merged = []
    seen_ids = set()
    # Pass 1: keep existing, but if a new contract has the same id, use the new one
    new_by_id = {c['contract_id']: c for c in new}
    for c in existing:
        cid = c['contract_id']
        if cid in new_by_id:
            merged.append(new_by_id[cid])  # file wins
        else:
            merged.append(c)
        seen_ids.add(cid)
    # Pass 2: append new contracts whose id wasn't in existing
    for c in new:
        if c['contract_id'] not in seen_ids:
            merged.append(c)
            seen_ids.add(c['contract_id'])
    return merged


def render_contracts_yaml(contracts):
    """Render 合同明细 list as YAML (with 2-space indent, list-of-dict)."""
    if not contracts:
        return ''
    lines = []
    for c in contracts:
        lines.append(f'  - 名称: {safe_yaml_str(c["template_name"])}')
        lines.append(f'    大类: {safe_yaml_str(c["category"])}')
        lines.append(f'    签约日期: {c["signed_at"]}')
        amt = c.get('signed_amount')
        if amt is not None:
            lines.append(f'    签约金额: {amt:.2f}' if isinstance(amt, float) else f'    签约金额: {amt}')
        lines.append(f'    合同状态: {safe_yaml_str(c["status"])}')
        lines.append(f'    合同编号: {safe_yaml_str(c["contract_id"])}')
    return '\n'.join(lines)


def render_main_categories(contracts):
    """List of unique 大类 values, ordered by first appearance."""
    seen = set()
    result = []
    for c in contracts:
        if c['category'] not in seen:
            seen.add(c['category'])
            result.append(c['category'])
    return result


# ── YAML field replace helpers ──────────────────────────────────────────

def replace_yaml_field(text, field, new_value_block):
    """Replace `field: ...` (single-line OR multi-line list) with new_value_block,
    operating ONLY inside the YAML frontmatter block (between the leading `---`
    fences). If the field is absent from frontmatter, insert it before the closing
    `---`. The markdown body is never touched — a body line that happens to begin
    with `field:` must not be mistaken for the frontmatter field (that bug silently
    corrupted notes and lost contract data).
    """
    fm = re.match(r'(?s)^(---\s*\n)(.*?)(\r?\n---)', text)
    if not fm:
        # No frontmatter block at the top — nothing safe to rewrite.
        return text
    head, fm_body, tail = fm.group(1), fm.group(2), fm.group(3)
    rest = text[fm.end():]

    # Match: "field: <value>" OR "field:\n  - ...\n  - ..." within the frontmatter.
    pattern = re.compile(
        rf'^{re.escape(field)}:.*?(?=^\S|\Z)',
        re.MULTILINE | re.DOTALL,
    )
    if pattern.search(fm_body):
        fm_body = pattern.sub(new_value_block.rstrip() + '\n', fm_body, count=1).rstrip('\n')
    else:
        # Insert at the end of the frontmatter body (the closing `\n---` follows).
        fm_body = fm_body.rstrip('\n') + '\n' + new_value_block.rstrip()

    return head + fm_body + tail + rest


def update_existing_student_yaml(md_path, customer_ids, signers, contracts):
    """Merge (or in --rewrite mode, replace) YAML frontmatter for an existing student.

    Merge mode (DEFAULT): reads existing 合同明细 from vault, merges with the input
    (multi-file union) by 合同编号 (file wins on conflict), PRESERVING vault contracts
    not present in any file. 合同 大类 union'd. 客户ID also merged with existing array.
    This is the safe behavior — a vault 档案 contract is never lost just because the
    latest export doesn't contain it.

    Rewrite mode (`--rewrite` flag): REPLACES 合同明细 / 合同 / 客户ID from the input,
    dropping any vault contract absent from every file. Use only to deliberately prune.
    """
    text = md_path.read_text(encoding='utf-8')

    # Defensive: never overwrite 私单 — even if upstream logic misroutes here.
    # 2026-05-17 bug: hardcoded PRIVATE_NO_SIGNING + accidental name collision
    # caused legacy sync-financial-to-vault.py to overwrite 王世杰's 私单 李想 with
    # 钟婷婷的 ERP 李想 数据. Belt-and-suspenders check.
    if _yaml_is_private(text):
        return False

    # In incremental mode, merge inputs with what's already in the vault file.
    if INCREMENTAL:
        existing_contracts = read_existing_contracts_from_yaml(md_path)
        contracts = merge_contracts_by_id(existing_contracts, contracts)
        # Also merge customer_ids with existing array
        m = re.search(r'^客户ID:\s*\[(.+?)\]', text, re.MULTILINE)
        if m:
            existing_cids = [c.strip() for c in m.group(1).split(',') if c.strip()]
            seen = set()
            merged_cids = []
            for c in existing_cids + list(customer_ids):
                if c not in seen:
                    merged_cids.append(c)
                    seen.add(c)
            customer_ids = merged_cids

    # Build new value blocks
    cid_block = '客户ID: [' + ', '.join(customer_ids) + ']'

    if signers and (set(signers) - {md_path.parent.name}):
        # Has alias signers (different from folder name)
        alias_list = sorted(s for s in signers if s != md_path.parent.name)
        if alias_list:
            signer_block = '合同签约人: [' + ', '.join(alias_list) + ']'
        else:
            signer_block = None
    else:
        signer_block = None

    cats = render_main_categories(contracts)
    cat_block = '合同: [' + ', '.join(cats) + ']' if cats else '合同: []'

    detail_block = '合同明细:\n' + render_contracts_yaml(contracts) if contracts else '合同明细: []'

    # Apply replacements
    new_text = text
    new_text = replace_yaml_field(new_text, '客户ID', cid_block)
    if signer_block is not None:
        new_text = replace_yaml_field(new_text, '合同签约人', signer_block)
    # Drop legacy 预收 / 已收 / 已交付 / 优惠 fields by rewriting 合同明细
    new_text = replace_yaml_field(new_text, '合同', cat_block)
    new_text = replace_yaml_field(new_text, '合同明细', detail_block)

    if new_text != text:
        if DRY_RUN:
            return True  # pretend changed, caller counts
        md_path.write_text(new_text, encoding='utf-8')
        return True
    return False


def create_new_student_folder(target_folder, advisor, intake_year, customer_id, contracts):
    """Build a new vault student folder + minimal YAML + subdirs."""
    folder = VAULT / target_folder
    if folder.exists():
        return False  # idempotent
    if DRY_RUN:
        return True  # pretend created
    folder.mkdir()
    (folder / '沟通记录').mkdir()
    (folder / '个性化材料').mkdir()
    (folder / '文书材料').mkdir()

    progress = progress_for_intake(intake_year)
    intake_label = f'{intake_year} fall' if intake_year else '待补'

    cats = render_main_categories(contracts)
    cat_str = '[' + ', '.join(cats) + ']' if cats else '[]'

    detail = render_contracts_yaml(contracts) if contracts else ''

    yaml_block = f"""---
标签: 学生档案
姓名: {target_folder}
当前进度: {progress}
入学年份: {intake_label}
目标地区: 待补
目前就读学校: 待补
专业:
意向专业方向:
中期顾问: {advisor}
前期顾问:
后期顾问:
客户邮箱:
最后沟通时间:
客户ID: [{customer_id}]
合同: {cat_str}
合同明细:
{detail}
---

# {target_folder} 的学生档案

## 备注

> 自动建档来源签约表（{TODAY}）— 中期顾问 = **{advisor}**
> ⚠ 该 folder 由 import-signings.py 自动创建。基本信息（学校 / 专业 / 邮箱）需补全。

## 沟通与纪要汇总
<!-- 使用 Agent 工作流每次更新时将链接追加到这里 -->

## 个性化材料

## 文书材料
"""
    (folder / f'{target_folder}.md').write_text(yaml_block, encoding='utf-8')
    return True


# ── Supabase upload ─────────────────────────────────────────────────────

def supabase_upsert_contracts(env, contract_rows):
    """Bulk UPSERT contract rows to Supabase. Uses PostgREST UPSERT
    (POST + Prefer: resolution=merge-duplicates)."""
    if DRY_RUN:
        print(f'   [DRY-RUN] would upsert {len(contract_rows)} contract rows to Supabase')
        return len(contract_rows)
    url = env.get('SUPABASE_URL')
    key = env.get('SUPABASE_SERVICE_ROLE_KEY')
    if not url or not key:
        print('⚠ Skipping Supabase upload (env missing)')
        return 0

    # Batch in chunks of 500
    BATCH = 500
    total_ok = 0
    failed_batches = []
    n_batches = (len(contract_rows) + BATCH - 1) // BATCH
    for i in range(0, len(contract_rows), BATCH):
        chunk = contract_rows[i:i + BATCH]
        bn = i // BATCH + 1
        body = json.dumps(chunk, ensure_ascii=False, default=str).encode('utf-8')
        # Retry each batch on transient network errors (e.g. a keep-alive drop —
        # "Remote end closed connection without response"). Previously a single
        # blip did `return total_ok`, aborting the loop and SILENTLY dropping
        # every remaining batch (8500/10643 pushed, 2143 lost on 2026-06-17).
        last_err = None
        for attempt in range(1, 4):
            req = urllib.request.Request(
                f'{url}/rest/v1/contracts',
                data=body,
                method='POST',
                headers={
                    'apikey': key,
                    'Authorization': f'Bearer {key}',
                    'Content-Type': 'application/json',
                    'Prefer': 'resolution=merge-duplicates,return=minimal',
                },
            )
            try:
                with urllib.request.urlopen(req, timeout=60) as resp:
                    if resp.status in (200, 201, 204):
                        total_ok += len(chunk)
                        print(f'   ✓ pushed batch {bn}/{n_batches}: {len(chunk)} rows')
                    else:
                        print(f'   ⚠ batch {bn} status {resp.status}: {resp.read()[:200]}')
                        failed_batches.append(bn)
                last_err = None
                break
            except urllib.error.HTTPError as e:
                # 4xx/5xx with a body = usually a data problem; retry won't help.
                last_err = f'HTTP {e.code}: {e.read()[:300]}'
                break
            except Exception as e:
                last_err = str(e)
                if attempt < 3:
                    print(f'   … batch {bn} attempt {attempt} failed ({e}); retrying')
                    time.sleep(2 * attempt)
        if last_err is not None:
            print(f'   ✗ batch {bn} failed after retries: {last_err}')
            failed_batches.append(bn)
    if failed_batches:
        print(f'   ⚠ {len(failed_batches)} batch(es) FAILED {failed_batches} — '
              f'{len(contract_rows) - total_ok} contracts NOT pushed. Re-run to retry.')
    return total_ok


def supabase_link_students(env, cid_to_folder, name_to_student_id):
    """For each customer_id with a folder, set students.customer_ids[] += that cid."""
    if DRY_RUN:
        n = sum(1 for f in {fld for fld in cid_to_folder.values()} if f in name_to_student_id)
        print(f'   [DRY-RUN] would PATCH {n} students.customer_ids')
        return n
    url = env.get('SUPABASE_URL')
    key = env.get('SUPABASE_SERVICE_ROLE_KEY')
    if not url or not key:
        return 0

    # folder -> [customer_ids]
    folder_to_cids = defaultdict(list)
    for cid, folder in cid_to_folder.items():
        folder_to_cids[folder].append(cid)

    n = 0
    for folder, cids in folder_to_cids.items():
        sid = name_to_student_id.get(folder)
        if sid is None:
            continue
        body = json.dumps({'customer_ids': cids}).encode('utf-8')
        req = urllib.request.Request(
            f'{url}/rest/v1/students?id=eq.{sid}',
            data=body,
            method='PATCH',
            headers={
                'apikey': key,
                'Authorization': f'Bearer {key}',
                'Content-Type': 'application/json',
                'Prefer': 'return=minimal',
            },
        )
        try:
            urllib.request.urlopen(req, timeout=15)
            n += 1
        except Exception as e:
            print(f'   ⚠ failed link student {folder}: {e}')
    return n


def supabase_backfill_contracts_student_id(env):
    """UPDATE contracts SET student_id via single Supabase RPC call.

    Migrated 2026-05-17 from per-cid PATCH loop (5000+ HTTP calls, ~20 min,
    susceptible to SSL EOF disconnects) to single RPC backfill_contracts_student_id()
    (1 call, <1 sec, transactional). See migration `backfill_contracts_student_id_rpc`.
    """
    if DRY_RUN:
        print(f'   [DRY-RUN] would call backfill_contracts_student_id() RPC')
        return 0
    url = env.get('SUPABASE_URL')
    key = env.get('SUPABASE_SERVICE_ROLE_KEY')
    if not url or not key:
        return 0
    req = urllib.request.Request(
        f'{url}/rest/v1/rpc/backfill_contracts_student_id',
        data=b'{}',
        method='POST',
        headers={
            'apikey': key,
            'Authorization': f'Bearer {key}',
            'Content-Type': 'application/json',
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            body = resp.read().decode('utf-8').strip()
            # PostgREST returns the scalar as a bare number
            return int(body) if body else 0
    except Exception as e:
        print(f'   ⚠ backfill RPC failed: {e}')
        return 0


def supabase_get_student_id_map(env):
    """Pull current students table → dict[name] = id."""
    url = env.get('SUPABASE_URL')
    key = env.get('SUPABASE_SERVICE_ROLE_KEY')
    if not url or not key:
        return {}
    req = urllib.request.Request(
        f'{url}/rest/v1/students?select=id,name&limit=1000',
        headers={'apikey': key, 'Authorization': f'Bearer {key}'},
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        rows = json.loads(resp.read())
    return {r['name']: r['id'] for r in rows}


# ── Main ────────────────────────────────────────────────────────────────

def main():
    if DRY_RUN:
        print('🔍 DRY-RUN MODE — no vault or Supabase writes will be performed\n')
    env = load_env()

    rows, COL = read_signings()
    vault_students = read_vault_students()
    vault_folder_names = {vs['name'] for vs in vault_students}
    print(f'📁 Vault: {len(vault_students)} folders')

    cid_to_folder, folder_to_cids, folder_signers = \
        build_customer_to_folder_map(rows, COL, vault_students)
    print(f'   {len(cid_to_folder)} customer_id mapped to {len(folder_to_cids)} vault folders')

    # ── Identify onboard candidates ──
    candidates = identify_onboard_candidates(rows, COL, cid_to_folder, vault_folder_names)
    print(f'\n🆕 Onboard candidates ({len(candidates)} new students for {ONBOARD_ADVISORS}):')

    # ── Group contracts by 客户ID for vault rewrite ──
    contracts_by_cid = defaultdict(list)
    for r in rows:
        cid = r[COL['客户ID']]
        contract_id = r[COL['合同编号']]
        signed = parse_date(r[COL['签约时间']])
        if not signed:
            continue
        contracts_by_cid[cid].append({
            'contract_id': contract_id,
            'customer_id': cid,
            'customer_name': r[COL['客户姓名']],
            'signed_at': signed,
            'signed_amount': r[COL['签约金额（实时）']] or 0,
            'template_name': r[COL['合同模板名称']] or '',
            'category': category(r[COL['合同模板名称']] or ''),
            'status': r[COL['合同状态']] or '',
            'is_application': is_yes(r[COL['是否包含留学申请']]),
            'is_training': is_yes(r[COL['是否包含语言培训']]),
            'is_research': is_yes(r[COL.get('是否科研类', -1)] if '是否科研类' in COL else None),
            'is_competition': is_yes(r[COL.get('是否竞赛类', -1)] if '是否竞赛类' in COL else None),
            'intake_year': r[COL['入学年']],
            'signing_advisor': r[COL.get('签约顾问', -1)] if '签约顾问' in COL else None,
            'early_advisor':   r[COL['前期顾问']] if r[COL['前期顾问']] else None,
            'mid_advisor':     r[COL['中期顾问']] if r[COL['中期顾问']] else None,
            'late_advisor':    r[COL['后期顾问']] if r[COL['后期顾问']] else None,
            'copywriter':      (r[COL['文案顾问']] if r[COL['文案顾问']] else None) if '文案顾问' in COL else None,
            'source_type':     r[COL.get('来源类型', -1)] if '来源类型' in COL else None,
            'channel_category': r[COL.get('渠道分类', -1)] if '渠道分类' in COL else None,
            'channel_detail':  r[COL.get('渠道明细', -1)] if '渠道明细' in COL else None,
            'referrer':        r[COL.get('推荐人', -1)] if '推荐人' in COL else None,
            'is_first_application': is_yes(r[COL.get('19财至查询截止日是否首次申请', -1)] if '19财至查询截止日是否首次申请' in COL else None),
            'is_first_company': is_yes(r[COL.get('是否当前公司首次签约', -1)] if '是否当前公司首次签约' in COL else None),
            'is_b2b':          is_yes(r[COL.get('是否B端', -1)] if '是否B端' in COL else None),
            'b2b_partner':     r[COL.get('B端合作方', -1)] if 'B端合作方' in COL else None,
            'city':            r[COL.get('城市', -1)] if '城市' in COL else None,
            'product_category': r[COL.get('产品类别', -1)] if '产品类别' in COL else None,
            'service_category': r[COL.get('主体服务类别', -1)] if '主体服务类别' in COL else None,
        })

    # Sort each customer's contracts by signed_at descending (newest first)
    for cid in contracts_by_cid:
        contracts_by_cid[cid].sort(key=lambda c: c['signed_at'], reverse=True)

    # ── Update existing vault students' YAML ──
    print(f'\n📝 Updating existing vault students:')
    yaml_updates = 0
    for vs in vault_students:
        cids = sorted(folder_to_cids.get(vs['name'], set()))
        if not cids:
            continue
        signers = sorted(folder_signers.get(vs['name'], set()))
        # Aggregate contracts across all cids for this folder
        agg = []
        for cid in cids:
            agg.extend(contracts_by_cid.get(cid, []))
        agg.sort(key=lambda c: c['signed_at'], reverse=True)
        if update_existing_student_yaml(vs['md_path'], cids, signers, agg):
            yaml_updates += 1
    print(f'   {yaml_updates} student YAMLs rewritten')

    # ── Onboard new vault folders ──
    print(f'\n🆕 Creating new vault folders:')
    onboarded = 0
    for c in candidates:
        target = c['target_folder']
        contracts = contracts_by_cid.get(c['cid'], [])
        if create_new_student_folder(target, c['advisor'], c['intake'], c['cid'], contracts):
            onboarded += 1
    print(f'   {onboarded} new folders created')

    # 新 onboard 的学生纳入"在库"集合 → 其合同也进推送白名单（见下方推送收窄）。
    for c in candidates:
        cid_to_folder.setdefault(c['cid'], c['target_folder'])

    # ── Push to Supabase contracts table ──
    print(f'\n📤 Push to Supabase contracts table:')
    db_rows = []
    for cid, contracts in contracts_by_cid.items():
        for c in contracts:
            db_rows.append(c)
    print(f'   {len(db_rows)} total contract rows to upsert')

    # Phase A (2026-05-17): vault YAML 是 advisor 归属的 source of truth.
    # 同一 cid 下所有 contract row 的 mid/early/late_advisor 一律按 vault YAML
    # 该 folder 的当前 中期/前期/后期顾问 写入 (多顾问取第一个 = primary).
    # vault 字段为空 → 保留 ERP 值 (避免无意删数据). vault 没有对应 folder → 不变.
    folder_to_advisors = {vs['name']: vs.get('advisors', {}) for vs in vault_students}

    override_n = 0
    for row in db_rows:
        folder = cid_to_folder.get(row['customer_id'])
        if not folder:
            continue
        adv = folder_to_advisors.get(folder, {})
        if adv.get('mid')   and row.get('mid_advisor')   != adv['mid']:
            row['mid_advisor']   = adv['mid'];   override_n += 1
        if adv.get('early') and row.get('early_advisor') != adv['early']:
            row['early_advisor'] = adv['early']
        if adv.get('late')  and row.get('late_advisor')  != adv['late']:
            row['late_advisor']  = adv['late']
    print(f'   🔁 vault override: {override_n} 行 mid_advisor 改写 (前期/后期 同步覆盖)')

    # ── 推送收窄 (2026-06-30) ──────────────────────────────────────────────
    # 只入「在库学生的全部合同」+「全公司非留学产品合同」。后者喂看板『月度产品节奏』
    # widget（2026-05-09 全公司口径，src/pages/internal/kanban/index.astro 的
    # rawMonthlyContracts 查询，is_application=false 无 student_id 过滤）。别部门的
    # 【留学】合同看板任何地方都不读 → 不推；DB 里的历史死行由 purge SQL 同口径清理。
    # 改这条口径要同步看 kanban 那个 widget 和 purge 谓词，三者必须一致。
    _before = len(db_rows)
    db_rows = [c for c in db_rows
               if c['customer_id'] in cid_to_folder or not c.get('is_application')]
    print(f'   🔎 推送收窄: {_before} → {len(db_rows)} 行'
          f'（在库 {len(cid_to_folder)} cid 全部合同 + 全公司非留学产品）')

    n_pushed = supabase_upsert_contracts(env, db_rows)
    print(f'   ✓ pushed {n_pushed} contracts')

    # ── Link students.customer_ids[] ──
    print(f'\n🔗 Link students.customer_ids[]:')
    name_to_id = supabase_get_student_id_map(env)
    n_linked = supabase_link_students(env, cid_to_folder, name_to_id)
    print(f'   ✓ linked {n_linked} students')

    # ── Backfill contracts.student_id by joining customer_id → students.customer_ids[] ──
    n_backfilled = supabase_backfill_contracts_student_id(env)
    print(f'   ✓ backfilled contracts.student_id for {n_backfilled} rows')

    # ── Summary ──
    print(f'\n═══ Summary ═══')
    print(f'  Signing rows read:           {len(rows)}')
    print(f'  Unique customer_ids:         {len(contracts_by_cid)}')
    print(f'  Vault students existing:     {len(vault_students)}')
    print(f'  Vault YAMLs rewritten:       {yaml_updates}')
    print(f'  New folders onboarded:       {onboarded}')
    print(f'  Contracts pushed Supabase:   {n_pushed}')
    print(f'  Students linked customer_ids:{n_linked}')
    if DRY_RUN:
        print(f'\n🔍 DRY-RUN COMPLETE — no vault or Supabase writes were performed.')
        print(f'   Re-run without --dry-run to apply.')
    else:
        print(f'\n📋 Next steps (if you want to fully refresh):')
        print(f'   /usr/local/bin/python3 scripts/import-erp-comms.py')
        print(f'   /opt/homebrew/bin/node scripts/sync-students-to-supabase.mjs')


if __name__ == '__main__':
    main()
