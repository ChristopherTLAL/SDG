#!/usr/bin/env python3
"""
ingest-qs-ranks.py — ONE-TIME / MANUAL: bake official QS World Rankings into canonical.

The official QS export (an .xlsx downloaded from qs.com) is treated as raw source material.
This script reads it ONCE and writes the numbers directly into the canonical dataset
(src/data/universities/universities.json) — the runtime NEVER reads the xlsx.

Matching is by normalised English name + country code, so name collisions across countries
are resolved correctly (e.g. "Newcastle University" UK #137 vs "The University of Newcastle"
Australia #227 — the merge bug this fixes).

Banded ranks (e.g. "791-800", "1001-1200", "1401+") are stored as the LOWER-BOUND INTEGER
(791, 1001, 1401) so existing numeric consumers (budget calculator sort, schools map) keep working.

Only qsRank is touched. Unmatched canonical rows keep their existing value.

Usage:
    python3 scripts/ingest-qs-ranks.py            # dry-run: report only
    python3 scripts/ingest-qs-ranks.py --apply    # write universities.json

After applying, run `node scripts/export-rankings.mjs` to sync the vault school-plan reference.
NOTE: build-universities.mjs seeds QS from the vault Top-200 file; if you ever re-run it,
re-run THIS script afterwards (this is the authoritative QS source, full ~1500 roster).
"""
import json, re, sys, os
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
CANON = os.path.join(HERE, '..', 'src', 'data', 'universities', 'universities.json')
VAULT = os.environ.get('OBSIDIAN_VAULT_ROOT', '/Users/shijie/Obsidian/规划看板')
XLSX = os.environ.get('QS_XLSX', os.path.join(VAULT, '附件', '2026 QS World University Rankings 1.3 (For qs.com).xlsx'))

# QS "Country/Territory" string -> our 2-letter country code (covers every code in canonical)
COUNTRY = {
    'United States of America': 'US', 'United Kingdom': 'UK', 'Australia': 'AU', 'Canada': 'CA',
    'Japan': 'JP', 'China (Mainland)': 'CN', 'Germany': 'DE', 'Netherlands': 'NL',
    'Republic of Korea': 'KR', 'Switzerland': 'CH', 'France': 'FR', 'Hong Kong SAR, China': 'HK',
    'Malaysia': 'MY', 'Sweden': 'SE', 'Belgium': 'BE', 'Denmark': 'DK', 'Spain': 'ES',
    'India': 'IN', 'Italy': 'IT', 'Saudi Arabia': 'SA', 'Taiwan': 'TW', 'Austria': 'AT',
    'Chile': 'CL', 'Finland': 'FI', 'Ireland': 'IE', 'Mexico': 'MX', 'New Zealand': 'NZ',
    'Singapore': 'SG', 'United Arab Emirates': 'AE', 'Argentina': 'AR', 'Brazil': 'BR',
    'Indonesia': 'ID', 'Kazakhstan': 'KZ', 'Norway': 'NO', 'Qatar': 'QA',
    'Russian Federation': 'RU', 'South Africa': 'ZA',
}

# Canonical name -> official QS spelling, for schools QS lists under a different name.
# Verified against the official QS2026 export.
ALIAS = {
    'Virginia Tech': 'Virginia Polytechnic Institute (Virginia Tech)',
    'University of Minnesota, Twin Cities': 'University of Minnesota (System)',
    'Binghamton University, State University of New York': 'Binghamton University SUNY',
    'University at Buffalo, State University of New York': 'University at Buffalo SUNY',
    'University of Tennessee, Knoxville': 'University of Tennessee',
    'College of William & Mary': 'William & Mary',
}
# Confirmed NOT in QS2026 (US News schools QS does not rank) — left null intentionally:
#   Texas Christian University, Pepperdine University, Villanova University, Gonzaga University


def norm(s):
    # Only strip 'the' — keep university/college/of so "Boston University" != "Boston College"
    # and "University of Miami" != "Miami University" (those collisions mis-assign ranks).
    s = (s or '').lower()
    s = re.sub(r'\(.*?\)', ' ', s)
    s = re.sub(r"[''`.,&\-–:]", ' ', s)
    s = re.sub(r'\bthe\b', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()

def lowbound(raw):
    """'137' -> 137, '791-800' -> 791, '1401+' -> 1401, '=92' -> 92, '' -> None"""
    s = str(raw).strip().lstrip('=')
    if '-' in s:
        s = s.split('-')[0]
    s = ''.join(ch for ch in s if ch.isdigit())
    return int(s) if s else None

def build_lookup():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb['QS']
    lookup = {}
    dups = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        rank, name, country = r[1], r[3], r[4]
        if not name:
            continue
        code = COUNTRY.get(str(country or '').strip())
        if not code:
            continue  # country not in our roster — skip
        lb = lowbound(rank)
        if lb is None:
            continue
        key = (norm(name), code)
        if key in lookup:
            dups.append((key, lookup[key], lb))
            if lb < lookup[key]:
                lookup[key] = lb  # keep the better (lower) rank on collision
        else:
            lookup[key] = lb
    return lookup, dups

def main():
    apply = '--apply' in sys.argv
    lookup, dups = build_lookup()
    data = json.load(open(CANON, encoding='utf-8'))

    filled, changed, unchanged, unmatched = [], [], 0, []
    for u in data:
        key = (norm(ALIAS.get(u['name'], u['name'])), u['country'])
        new = lookup.get(key)
        if new is None:
            unmatched.append((u['country'], u['name']))
            continue
        old = u.get('qsRank')
        if old in (None, ''):
            filled.append((u['country'], u['nameCn'], u['name'], new))
            if apply: u['qsRank'] = new
        elif old != new:
            changed.append((u['country'], u['nameCn'], u['name'], old, new))
            if apply: u['qsRank'] = new
        else:
            unchanged += 1

    print(f"QS lookup entries: {len(lookup)} | dup keys: {len(dups)}")
    print(f"canonical rows: {len(data)}")
    print(f"  filled (was null):   {len(filled)}")
    print(f"  changed (corrected): {len(changed)}")
    print(f"  unchanged (match):   {unchanged}")
    print(f"  unmatched (kept):    {len(unmatched)}")
    print(f"\n--- CHANGED existing values (review!) ---")
    for c, cn, en, old, new in sorted(changed, key=lambda x: abs((x[4] or 0)-(x[3] or 0)), reverse=True):
        print(f"  [{c}] {cn} ({en}): {old} -> {new}")
    print(f"\n--- FILLED (sample 30 of {len(filled)}) ---")
    for c, cn, en, new in sorted(filled, key=lambda x: x[3])[:30]:
        print(f"  [{c}] QS {new:>5}  {cn} ({en})")
    print(f"\n--- UNMATCHED (sample 25 of {len(unmatched)}) ---")
    for c, en in unmatched[:25]:
        print(f"  [{c}] {en}")

    if apply:
        json.dump(data, open(CANON, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
        open(CANON, 'a', encoding='utf-8').write('\n')
        print(f"\n✅ WROTE {CANON}")
    else:
        print(f"\n(dry-run — pass --apply to write)")

if __name__ == '__main__':
    main()
