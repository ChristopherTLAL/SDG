#!/usr/bin/env python3
"""
Parse upstream PhD-skill Excel outputs into a unified supervisor list.

Auto-discovers the latest `*Supervisor_Shortlist_*.xlsx` (from vault skill
`phd_supervisor_match`) and `*Faculty_Research_*.xlsx` (from vault skill
`faculty-deep-dive`) in `01_Student/{name}/个性化材料/`, forward-fills merged
cells, joins both sources by professor name (fuzzy), respects "套磁过了"
annotations, and emits a unified JSON or markdown table.

Used by the `phd-cold-mail` skill (Phase 0) so the user does not have to
manually convert Excel rows into the skill's markdown-table input format.

Usage:
    python3 parse_shortlist.py STUDENT_NAME [--vault PATH] [--markdown]
    python3 parse_shortlist.py --shortlist PATH.xlsx [--faculty PATH.xlsx] [--markdown]

Output: JSON on stdout (default) or a markdown table (--markdown). Diagnostic
lines go to stderr.
"""

import argparse
import json
import os
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(2)


# Annotations the user adds in-Excel to mark a professor as "already contacted"
SKIP_PATTERNS = re.compile(
    r"(套磁过了|套磁完成|套磁成功|已联系|已套磁|已发出|已发送|已发邮件|reply received|已回复|contacted)",
    re.I,
)


def find_latest(folder: Path, pattern: str):
    """Find the latest Excel matching pattern (by YYMMDD suffix, else mtime)."""
    if not folder.exists():
        return None
    candidates = list(folder.glob(pattern))
    if not candidates:
        return None

    def date_key(p: Path) -> str:
        m = re.search(r"_(\d{6})\.xlsx$", p.name)
        if m:
            return m.group(1)
        return f"00{int(p.stat().st_mtime)}"

    return sorted(candidates, key=date_key, reverse=True)[0]


# Header aliases — resolves the column index by name so we tolerate user-edited
# Excels that add extra columns (e.g., `#` sequence, `Top20` rank) or rename
# headers slightly. Maps lowercase trimmed header → canonical field.
HEADER_ALIASES = {
    # region / 地区
    "地区": "region", "region": "region", "country": "region",
    # school / 学校
    "学校": "school", "school": "school", "university": "school", "机构": "school",
    # department / 系
    "院系/方向": "department", "院系": "department", "系": "department",
    "department": "department", "dept": "department", "方向": "department",
    # supervisor name / 导师姓名
    "导师姓名": "name", "导师": "name", "姓名": "name",
    "supervisor name": "name", "supervisor": "name", "name": "name", "professor": "name",
    # match reason / 匹配原因
    "匹配原因": "match_reason", "match reason": "match_reason", "reason": "match_reason",
    "总结": "match_reason", "summary": "match_reason",
    # homepage / 主页链接
    "主页链接": "homepage", "homepage": "homepage", "url": "homepage", "link": "homepage",
    # Top N rank (★1 ... ★20 etc.)
    "top20": "top_rank", "top": "top_rank", "rank": "top_rank", "排名": "top_rank",
    "top10": "top_rank", "top30": "top_rank",
}


def _resolve_headers(ws):
    """Map header-row strings → canonical field names → column index (1-based)."""
    col_map = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if not isinstance(h, str):
            continue
        key = h.strip().lower()
        if key in HEADER_ALIASES:
            col_map.setdefault(HEADER_ALIASES[key], c)
    return col_map


def parse_shortlist_xlsx(path: Path):
    """Parse a supervisor shortlist Excel.

    Tolerates two layouts:
    - Original `phd_supervisor_match` output: 6 cols [地区, 学校, 院系/方向, 导师姓名, 匹配原因, 主页链接]
    - User-edited variant: 8 cols [#, 地区, 学校, 系, 导师姓名, 匹配原因, 主页链接, Top20]

    Column positions are resolved by header NAME via `HEADER_ALIASES` — adding
    new columns or renaming should not break parsing.

    Region/School/Department are typically merged across professor groups; this
    parser forward-fills them based on column position discovered from headers.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    col = _resolve_headers(ws)

    required = ("name", "school")
    missing = [f for f in required if f not in col]
    if missing:
        raise ValueError(
            f"{path.name}: required columns missing (resolved: {col}; missing: {missing}). "
            "Headers must include 导师姓名 + 学校 (or English equivalents)."
        )

    rows = []
    region = uni = dept = None
    for r in range(2, ws.max_row + 1):
        def get(field):
            c_idx = col.get(field)
            if not c_idx:
                return None
            return ws.cell(r, c_idx).value

        # Forward-fill the merged columns
        a = get("region")
        if isinstance(a, str) and a.strip():
            region = a.strip()
        b = get("school")
        if isinstance(b, str) and b.strip():
            uni = b.strip()
        c_val = get("department")
        if isinstance(c_val, str) and c_val.strip():
            dept = c_val.strip()

        d = get("name")
        if not (isinstance(d, str) and d.strip()):
            continue
        name = d.strip()

        e = get("match_reason")
        match_reason = e.strip() if isinstance(e, str) else ""
        f = get("homepage")
        homepage = f.strip() if isinstance(f, str) else ""
        g = get("top_rank")
        # ★12 → 12, "Top 5" → 5, plain int → int, else None
        top_rank = None
        if g is not None:
            m = re.search(r"\d+", str(g))
            if m:
                top_rank = int(m.group(0))

        skip = bool(SKIP_PATTERNS.search(match_reason))
        clean_reason = SKIP_PATTERNS.sub("", match_reason).strip()
        rows.append(
            {
                "source": "shortlist",
                "name": name,
                "region": region,
                "school": uni,
                "department": dept,
                "match_reason": clean_reason,
                "homepage": homepage,
                "top_rank": top_rank,
                "skip": skip,
                "skip_note": match_reason if skip else None,
            }
        )
    return rows


def parse_faculty_xlsx(path: Path):
    """Parse `faculty-deep-dive` Excel.

    Layout: A=导师姓名, B=学校, C=院系/方向, D=导师背景 (all merged across N
    paper rows), E=论文 (APA, one per row), F=论文解读 (one per row).
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    current = None
    for r in range(2, ws.max_row + 1):
        a, b, c, d, e, f = (ws.cell(r, col).value for col in range(1, 7))
        if isinstance(a, str) and a.strip():
            if current:
                rows.append(current)
            current = {
                "source": "faculty",
                "name": a.strip(),
                "school": (b.strip() if isinstance(b, str) else ""),
                "department": (c.strip() if isinstance(c, str) else ""),
                "background": (d.strip() if isinstance(d, str) else ""),
                "papers": [],
                "skip": False,
            }
        if current and isinstance(e, str) and e.strip():
            current["papers"].append(
                {
                    "citation": e.strip(),
                    "explanation": (f.strip() if isinstance(f, str) else ""),
                }
            )
    if current:
        rows.append(current)
    return rows


def normalize_name(name: str) -> str:
    """Normalize professor name for fuzzy matching across the two Excels.

    Handles "Prof. Matt Kusner" ↔ "Matt Kusner" ↔ "Prof Matt Kusner".
    """
    n = name.strip().rstrip(",")
    n = re.sub(
        r"^(Prof\.?|Professor|Dr\.?|Associate\s+Prof\.?|Asst\.?\s+Prof\.?|Mr\.?|Ms\.?|Mrs\.?)\s+",
        "",
        n,
        flags=re.I,
    )
    n = re.sub(r"\s+", " ", n)
    return n.lower()


def merge(shortlist, faculty):
    """Join shortlist + faculty rows by fuzzy professor name."""
    by_name = {}
    order = []
    for s in shortlist:
        key = normalize_name(s["name"])
        if key not in by_name:
            order.append(key)
        by_name[key] = dict(s)
    for f in faculty:
        key = normalize_name(f["name"])
        if key in by_name:
            row = by_name[key]
            row["background"] = f.get("background")
            row["papers"] = f.get("papers")
            row["source"] = row["source"] + "+faculty"
        else:
            row = dict(f)
            order.append(key)
            by_name[key] = row
    return [by_name[k] for k in order]


def to_markdown(rows):
    """Render the unified list as a markdown table for human review.

    Schema matches the `phd-cold-mail` SKILL.md `导师名单.md` shape so the
    output can be saved as-is if the user prefers a markdown intermediate.
    """
    headers = ["★", "Supervisor Name", "机构/学校", "院系", "邮箱", "总结", "重要论文", "重要链接", "_状态"]
    lines = ["| " + " | ".join(headers) + " |", "|" + "|".join(["---"] * len(headers)) + "|"]
    for r in rows:
        papers = r.get("papers") or []
        paper_cell = "<br>".join(p["citation"] for p in papers) if papers else ""
        summary = r.get("background") or r.get("match_reason", "") or ""
        homepage = (r.get("homepage") or "").replace("\n", "<br>")
        rank = r.get("top_rank")
        rank_cell = f"★{rank}" if rank else ""
        cells = [
            rank_cell,
            r.get("name", ""),
            r.get("school", ""),
            r.get("department", "").split("\n")[0],
            "",  # 邮箱 — research subagent fills in
            summary.replace("\n", " ").replace("|", "\\|"),
            paper_cell.replace("|", "\\|"),
            homepage.replace("|", "\\|"),
            "已联系" if r.get("skip") else "",
        ]
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("student", nargs="?", help="Student folder name in vault `01_Student/`")
    ap.add_argument(
        "--vault",
        default=os.environ.get("OBSIDIAN_VAULT_ROOT", "/Users/shijie/Obsidian/规划看板"),
        help="Vault root (default: $OBSIDIAN_VAULT_ROOT or /Users/shijie/Obsidian/规划看板)",
    )
    ap.add_argument("--shortlist", help="Explicit path to Supervisor_Shortlist xlsx")
    ap.add_argument("--faculty", help="Explicit path to Faculty_Research xlsx")
    ap.add_argument("--markdown", action="store_true", help="Emit markdown table instead of JSON")
    ap.add_argument(
        "--top-only",
        type=int,
        metavar="N",
        help="Only keep rows with top_rank <= N (drops un-ranked rows). Useful for the first batch.",
    )
    ap.add_argument(
        "--include-skipped",
        action="store_true",
        help="Include rows marked as already contacted (default: drop them).",
    )
    args = ap.parse_args()

    shortlist_path = Path(args.shortlist) if args.shortlist else None
    faculty_path = Path(args.faculty) if args.faculty else None
    folder = None
    if args.student:
        folder = Path(args.vault) / "01_Student" / args.student / "个性化材料"
        if shortlist_path is None:
            shortlist_path = find_latest(folder, "*Supervisor_Shortlist_*.xlsx")
        if faculty_path is None:
            faculty_path = find_latest(folder, "*Faculty_Research_*.xlsx")

    if not shortlist_path and not faculty_path:
        print(f"ERROR: No upstream Excel found. Looked in: {folder}", file=sys.stderr)
        sys.exit(1)

    shortlist = parse_shortlist_xlsx(shortlist_path) if shortlist_path else []
    faculty = parse_faculty_xlsx(faculty_path) if faculty_path else []
    merged = merge(shortlist, faculty)

    total = len(merged)
    skipped = sum(1 for r in merged if r.get("skip"))

    # Apply filters
    filtered = list(merged)
    if not args.include_skipped:
        filtered = [r for r in filtered if not r.get("skip")]
    if args.top_only is not None:
        filtered = [r for r in filtered if r.get("top_rank") and r["top_rank"] <= args.top_only]
        # Sort by rank for stable batch order
        filtered.sort(key=lambda r: r["top_rank"])

    print(
        f"# Sources:\n#   shortlist: {shortlist_path}\n#   faculty:   {faculty_path}\n"
        f"# Parsed: {len(shortlist)} shortlist + {len(faculty)} faculty → "
        f"{total} unique professors ({skipped} already contacted)\n"
        f"# Filters: top_only={args.top_only}, include_skipped={args.include_skipped} → {len(filtered)} kept",
        file=sys.stderr,
    )

    if args.markdown:
        print(to_markdown(filtered))
    else:
        json.dump(filtered, sys.stdout, ensure_ascii=False, indent=2)
        print()


if __name__ == "__main__":
    main()
